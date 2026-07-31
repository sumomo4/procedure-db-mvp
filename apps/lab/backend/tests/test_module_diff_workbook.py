"""Tests for layout-preserving module diff workbooks."""

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image as PillowImage

from app.core.module_diff_workbook import DIFF_FILL_COLOR, build_module_diff_workbook
from app.core.responses import (
    ModuleCreateDeviceHeaderInput,
    ModuleCreateRequest,
    ModuleCreateRowDeviceEntryInput,
    ModuleCreateRowImageInput,
    ModuleCreateRowInput,
    ModuleDiffData,
    ModuleDiffRowData,
    ModuleDiffSummaryData,
    ModuleRowData,
    ModuleRowDeviceEntryData,
    ModuleRowImageData,
)


def _build_row(
    module_row_id: int,
    row_order: int,
    work_text: str,
    command_text: str,
) -> ModuleRowData:
    """Build one compact procedure row for workbook assertions."""

    return ModuleRowData(
        module_row_id=module_row_id,
        row_order=row_order,
        row_type="step",
        major_no="1",
        middle_no="1",
        minor_no=str(row_order),
        tech_doc_text="操作手順",
        work_text=work_text,
        indent_level=0,
        expected_result="正常終了すること",
        time_text=None,
        window_text=None,
        p_text=None,
        command_text=command_text,
        note=None,
        device_entries=[
            ModuleRowDeviceEntryData(
                slot_no=1,
                time_text=None,
                window_text="TT",
                p_text="#",
                command_text=command_text,
            )
        ],
        images=[],
    )


def _build_import_workbook() -> bytes:
    """Build an imported module workbook with layout details to preserve."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "初期点検手順"
    worksheet.merge_cells("A1:I1")
    worksheet["A1"] = "初期点検手順"
    worksheet["A1"].font = Font(bold=True, size=15)
    worksheet["A4"] = "通常"
    for column_index, label in enumerate(
        ("大", "中", "小", "技術資料名", "作業内容", "", "", "", "確認事項 or 項目"),
        start=1,
    ):
        worksheet.cell(row=5, column=column_index, value=label)
    worksheet["J5"] = "時刻"
    worksheet["K5"] = "window"
    worksheet["L5"] = "P"
    worksheet["M5"] = "コマンド"

    worksheet["A6"] = "1"
    worksheet["B6"] = "1"
    worksheet["C6"] = "1"
    worksheet["D6"] = "操作手順"
    worksheet["E6"] = "設定値を確認する"
    worksheet["I6"] = "正常終了すること"
    worksheet["K6"] = "TT"
    worksheet["L6"] = "#"
    worksheet["M6"] = "show startup-config"

    worksheet["A7"] = "1"
    worksheet["B7"] = "2"
    worksheet["C7"] = "1"
    worksheet["D7"] = "操作手順"
    worksheet["E7"] = "ログを保存する"
    worksheet["I7"] = "保存されること"
    worksheet["K7"] = "TT"
    worksheet["L7"] = "#"
    worksheet["M7"] = "save log"

    worksheet.column_dimensions["E"].width = 44
    worksheet.row_dimensions[6].height = 36
    worksheet["G6"].fill = PatternFill(fill_type="solid", fgColor="FFF4E4C1")

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_import_payload() -> ModuleCreateRequest:
    """Return normalized data corresponding to the imported workbook."""

    return ModuleCreateRequest(
        module_name="初期点検手順",
        source_xlsx_path="module.xlsx",
        device_headers=[ModuleCreateDeviceHeaderInput(slot_no=1)],
        rows=[
            ModuleCreateRowInput(
                row_order=1,
                row_type="step",
                major_no="1",
                middle_no="1",
                minor_no="1",
                tech_doc_text="操作手順",
                work_text="設定値を確認する",
                indent_level=0,
                expected_result="正常終了すること",
                device_entries=[
                    ModuleCreateRowDeviceEntryInput(
                        slot_no=1,
                        window_text="TT",
                        p_text="#",
                        command_text="show startup-config",
                    )
                ],
            ),
            ModuleCreateRowInput(
                row_order=2,
                row_type="step",
                major_no="1",
                middle_no="2",
                minor_no="1",
                tech_doc_text="操作手順",
                work_text="ログを保存する",
                indent_level=0,
                expected_result="保存されること",
                device_entries=[
                    ModuleCreateRowDeviceEntryInput(
                        slot_no=1,
                        window_text="TT",
                        p_text="#",
                        command_text="save log",
                    )
                ],
            ),
        ],
    )


def test_build_module_diff_workbook_preserves_layout_and_highlights_cells() -> None:
    """Both sheets should retain the import layout and highlight only differences."""

    before = _build_row(10, 1, "設定を確認する", "show running-config")
    after = _build_row(-1, 1, "設定値を確認する", "show startup-config")
    added = _build_row(-2, 2, "ログを保存する", "save log")
    diff = ModuleDiffData(
        module_id=1,
        module_key="MOD-001",
        module_name="初期点検手順",
        from_version=1,
        to_version=2,
        summary=ModuleDiffSummaryData(
            added_count=1,
            removed_count=0,
            changed_count=1,
            unchanged_count=0,
        ),
        rows=[
            ModuleDiffRowData(
                status="changed",
                row_key="row_order:1->1",
                before=before,
                after=after,
                changed_fields=["work_text", "device_entries"],
                similarity=0.8,
            ),
            ModuleDiffRowData(
                status="added",
                row_key="added:2",
                before=None,
                after=added,
                changed_fields=[],
                similarity=None,
            ),
        ],
    )

    workbook = load_workbook(
        BytesIO(
            build_module_diff_workbook(
                diff,
                _build_import_workbook(),
                _build_import_payload(),
            )
        )
    )

    assert workbook.sheetnames == ["比較元", "比較先"]
    source_sheet = workbook["比較元"]
    target_sheet = workbook["比較先"]
    expected_color = f"FF{DIFF_FILL_COLOR}"

    assert "A1:I1" in source_sheet.merged_cells
    assert "A1:I1" in target_sheet.merged_cells
    assert source_sheet.column_dimensions["E"].width == 44
    assert target_sheet.column_dimensions["E"].width == 44
    assert source_sheet.row_dimensions[6].height == 36
    assert target_sheet.row_dimensions[6].height == 36
    assert source_sheet["A1"].font.bold is True
    assert target_sheet["A1"].font.bold is True

    assert source_sheet["E6"].value == "設定を確認する"
    assert target_sheet["E6"].value == "設定値を確認する"
    assert source_sheet["E6"].fill.fgColor.rgb == expected_color
    assert target_sheet["E6"].fill.fgColor.rgb == expected_color
    assert source_sheet["M6"].value == "show running-config"
    assert target_sheet["M6"].value == "show startup-config"
    assert source_sheet["M6"].fill.fgColor.rgb == expected_color
    assert target_sheet["M6"].fill.fgColor.rgb == expected_color
    assert source_sheet["G6"].fill.fgColor.rgb == "FFF4E4C1"
    assert target_sheet["G6"].fill.fgColor.rgb == "FFF4E4C1"

    assert source_sheet["A7"].value is None
    assert target_sheet["A7"].value == "1"
    assert source_sheet["A7"].fill.fgColor.rgb != expected_color
    assert target_sheet["A7"].fill.fgColor.rgb == expected_color


def _build_image_workbook() -> bytes:
    """Build an imported workbook with one tall image row."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "画像モジュール"
    for column_index, label in enumerate(
        ("大", "中", "小", "技術資料名", "作業内容", "", "", "", "確認事項 or 項目"),
        start=1,
    ):
        worksheet.cell(row=5, column=column_index, value=label)
    worksheet["J5"] = "時刻"
    worksheet["K5"] = "window"
    worksheet["L5"] = "P"
    worksheet["M5"] = "コマンド"
    worksheet["I6"] = "画像を確認すること"
    worksheet.row_dimensions[6].height = 229.5
    worksheet.column_dimensions["E"].width = 24
    worksheet.column_dimensions["F"].width = 12
    worksheet.column_dimensions["G"].width = 12
    worksheet.column_dimensions["H"].width = 24

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _create_test_image(path: Path, color: str) -> None:
    """Create one deliberately oversized image."""

    PillowImage.new("RGB", (1000, 1200), color=color).save(path)


def test_image_difference_is_not_colored_and_images_fit_row(tmp_path: Path) -> None:
    """Image-only metadata changes should not color cells or overflow the row."""

    before_image_path = tmp_path / "before.png"
    after_image_path = tmp_path / "after.png"
    _create_test_image(before_image_path, "navy")
    _create_test_image(after_image_path, "teal")

    before = ModuleRowData(
        module_row_id=20,
        row_order=1,
        row_type="step",
        major_no=None,
        middle_no=None,
        minor_no=None,
        tech_doc_text=None,
        work_text=None,
        indent_level=0,
        expected_result="画像を確認すること",
        time_text=None,
        window_text=None,
        p_text=None,
        command_text=None,
        note=None,
        device_entries=[],
        images=[
            ModuleRowImageData(
                module_row_image_id=1,
                image_key="before-image",
                image_path=str(before_image_path),
                anchor_cell="E6",
                width_px=1000,
                height_px=1200,
            )
        ],
    )
    after = before.model_copy(
        update={
            "module_row_id": -1,
            "indent_level": None,
            "images": [
                ModuleRowImageData(
                    module_row_image_id=-1,
                    image_key="after-image",
                    image_path=str(after_image_path),
                    anchor_cell="E6",
                    width_px=1000,
                    height_px=1200,
                )
            ],
        }
    )
    diff = ModuleDiffData(
        module_id=1,
        module_key="MOD-001",
        module_name="画像モジュール",
        from_version=1,
        to_version=2,
        summary=ModuleDiffSummaryData(changed_count=1),
        rows=[
            ModuleDiffRowData(
                status="changed",
                row_key="row_order:1->1",
                before=before,
                after=after,
                changed_fields=["indent_level", "images"],
                similarity=1.0,
            )
        ],
    )
    payload = ModuleCreateRequest(
        module_name="画像モジュール",
        device_headers=[ModuleCreateDeviceHeaderInput(slot_no=1)],
        rows=[
            ModuleCreateRowInput(
                row_order=1,
                row_type="step",
                expected_result="画像を確認すること",
                images=[
                    ModuleCreateRowImageInput(
                        image_key="after-image",
                        image_path=str(after_image_path),
                        anchor_cell="E6",
                        width_px=1000,
                        height_px=1200,
                    )
                ],
            )
        ],
    )

    workbook = load_workbook(
        BytesIO(build_module_diff_workbook(diff, _build_image_workbook(), payload))
    )
    expected_color = f"FF{DIFF_FILL_COLOR}"
    for worksheet in workbook.worksheets:
        assert worksheet["E6"].fill.fgColor.rgb != expected_color
        assert worksheet["I6"].fill.fgColor.rgb != expected_color
        assert len(worksheet._images) == 1
        image = worksheet._images[0]
        assert round(image.anchor.ext.cy / 9525) <= 294
        assert round(image.anchor.ext.cx / 9525) <= 245
