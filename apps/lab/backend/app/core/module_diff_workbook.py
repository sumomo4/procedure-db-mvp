"""Build module diff workbooks while preserving the imported Excel layout."""

from __future__ import annotations

from copy import copy
from io import BytesIO
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.core.responses import (
    ModuleCreateRequest,
    ModuleDiffData,
    ModuleDiffRowData,
    ModuleRowData,
)


DIFF_FILL_COLOR = "E4D7F5"
_DIFF_FILL = PatternFill(fill_type="solid", fgColor=f"FF{DIFF_FILL_COLOR}")
_WORK_START_COLUMN = 5
_WORK_END_COLUMN = 8
_EXPECTED_RESULT_COLUMN = 9
_DEVICE_START_COLUMN = 10
_DEVICE_COLUMN_COUNT = 4
_HEADER_LABEL_SETS = (
    {"大", "中", "小", "作業内容"},
    {"major", "middle", "minor", "work"},
)


def _normalize_label(value: object | None) -> str:
    """Normalize one worksheet label for loose header matching."""

    if value is None:
        return ""
    return "".join(str(value).lower().split())


def _find_header_row(worksheet: Worksheet) -> int:
    """Find the procedure table header row in the imported worksheet."""

    for row_index in range(1, worksheet.max_row + 1):
        values = {
            _normalize_label(worksheet.cell(row=row_index, column=column_index).value)
            for column_index in range(1, min(worksheet.max_column, 40) + 1)
        }
        if any(
            all(_normalize_label(label) in values for label in required_labels)
            for required_labels in _HEADER_LABEL_SETS
        ):
            return row_index
    raise ValueError("Worksheet header row was not found.")


def _image_anchor_row(image: ExcelImage) -> int | None:
    """Return the one-based row of an openpyxl worksheet image."""

    anchor = image.anchor
    if isinstance(anchor, str):
        match = re.search(r"\d+", anchor)
        return int(match.group()) if match is not None else None

    marker = getattr(anchor, "_from", None)
    row_index = getattr(marker, "row", None)
    return int(row_index) + 1 if row_index is not None else None


def _device_slot_count(payload: ModuleCreateRequest) -> int:
    """Return the number of device column blocks represented by the import."""

    slot_nos = {header.slot_no for header in payload.device_headers}
    slot_nos.update(
        entry.slot_no
        for row in payload.rows
        for entry in row.device_entries
    )
    return max(slot_nos, default=1)


def _tracked_last_column(payload: ModuleCreateRequest) -> int:
    """Return the final module data column for the imported workbook."""

    return _DEVICE_START_COLUMN + (_device_slot_count(payload) * _DEVICE_COLUMN_COUNT) - 1


def _source_row_indices(
    worksheet: Worksheet,
    payload: ModuleCreateRequest,
    header_row: int,
) -> list[int]:
    """Map normalized module rows back to physical Excel row numbers."""

    last_column = _tracked_last_column(payload)
    image_rows = {
        row_index
        for image in worksheet._images
        if (row_index := _image_anchor_row(image)) is not None
    }
    image_rows.update(
        int(match.group())
        for row in payload.rows
        for image in row.images
        if (match := re.search(r"\d+", image.anchor_cell)) is not None
    )

    row_indices = [
        row_index
        for row_index in range(header_row + 1, worksheet.max_row + 1)
        if row_index in image_rows
        or any(
            worksheet.cell(row=row_index, column=column_index).value not in (None, "")
            for column_index in range(1, last_column + 1)
        )
    ]

    if len(row_indices) >= len(payload.rows):
        return row_indices[: len(payload.rows)]

    next_row = row_indices[-1] + 1 if row_indices else header_row + 1
    while len(row_indices) < len(payload.rows):
        row_indices.append(next_row)
        next_row += 1
    return row_indices


def _excel_row_for_order(row_indices: list[int], row_order: int) -> int:
    """Resolve a normalized one-based row order to a physical Excel row."""

    if row_order <= len(row_indices):
        return row_indices[row_order - 1]
    return row_indices[-1] + row_order - len(row_indices)


def _copy_row_style(worksheet: Worksheet, source_row: int, target_row: int, last_column: int) -> None:
    """Copy row dimensions and cell styles for an appended comparison row."""

    source_dimension = worksheet.row_dimensions[source_row]
    target_dimension = worksheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel

    for column_index in range(1, last_column + 1):
        source_cell = worksheet.cell(row=source_row, column=column_index)
        target_cell = worksheet.cell(row=target_row, column=column_index)
        if not isinstance(source_cell, MergedCell):
            target_cell._style = copy(source_cell._style)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)


def _ensure_rows(
    worksheet: Worksheet,
    row_indices: list[int],
    required_count: int,
    last_column: int,
) -> list[int]:
    """Append styled physical rows when the reference has more data rows."""

    if required_count <= len(row_indices):
        return row_indices

    template_row = row_indices[-1]
    next_row = template_row + 1
    while len(row_indices) < required_count:
        _copy_row_style(worksheet, template_row, next_row, last_column)
        row_indices.append(next_row)
        next_row += 1
    return row_indices


def _clear_data_cells(worksheet: Worksheet, row_indices: Iterable[int], last_column: int) -> None:
    """Clear values from the comparison-source copy without altering styles."""

    for row_index in row_indices:
        for column_index in range(1, last_column + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if not isinstance(cell, MergedCell):
                cell.value = None


def _entry_map(row: ModuleRowData) -> dict[int, tuple[str | None, str | None, str | None, str | None]]:
    """Return device values keyed by slot number."""

    entries = {
        entry.slot_no: (
            entry.time_text,
            entry.window_text,
            entry.p_text,
            entry.command_text,
        )
        for entry in row.device_entries
    }
    if 1 not in entries:
        entries[1] = (
            row.time_text,
            row.window_text,
            row.p_text,
            row.command_text,
        )
    return entries


def _write_reference_row(worksheet: Worksheet, excel_row: int, row: ModuleRowData, slot_count: int) -> None:
    """Restore one persisted module row into the source-layout copy."""

    values = (
        row.major_no,
        row.middle_no,
        row.minor_no,
        row.tech_doc_text,
    )
    for column_index, value in enumerate(values, start=1):
        worksheet.cell(row=excel_row, column=column_index).value = value

    for column_index in range(_WORK_START_COLUMN, _WORK_END_COLUMN + 1):
        worksheet.cell(row=excel_row, column=column_index).value = None
    work_column = min(
        _WORK_START_COLUMN + (row.indent_level or 0),
        _WORK_END_COLUMN,
    )
    worksheet.cell(row=excel_row, column=work_column).value = row.work_text
    worksheet.cell(row=excel_row, column=_EXPECTED_RESULT_COLUMN).value = row.expected_result

    entries = _entry_map(row)
    for slot_no in range(1, slot_count + 1):
        start_column = _DEVICE_START_COLUMN + ((slot_no - 1) * _DEVICE_COLUMN_COUNT)
        entry_values = entries.get(slot_no, (None, None, None, None))
        for offset, value in enumerate(entry_values):
            worksheet.cell(row=excel_row, column=start_column + offset).value = value


def _add_reference_images(worksheet: Worksheet, diff: ModuleDiffData) -> None:
    """Add persisted module images to the comparison-source sheet."""

    worksheet._images = []
    added_keys: set[str] = set()
    for diff_row in diff.rows:
        row = diff_row.before
        if row is None:
            continue
        for image in row.images:
            if image.image_key in added_keys:
                continue
            image_path = Path(image.image_path)
            if not image_path.is_file():
                continue
            excel_image = ExcelImage(image_path)
            _fit_image_to_cell_area(
                worksheet,
                excel_image,
                image.anchor_cell,
                image.width_px,
                image.height_px,
            )
            worksheet.add_image(excel_image)
            added_keys.add(image.image_key)


def _add_incoming_images(worksheet: Worksheet, payload: ModuleCreateRequest) -> None:
    """Restore imported images to the comparison-target sheet at a safe size."""

    worksheet._images = []
    added_keys: set[str] = set()
    for row in payload.rows:
        for image in row.images:
            if image.image_key in added_keys:
                continue
            image_path = Path(image.image_path)
            if not image_path.is_file():
                continue
            excel_image = ExcelImage(image_path)
            _fit_image_to_cell_area(
                worksheet,
                excel_image,
                image.anchor_cell,
                image.width_px,
                image.height_px,
            )
            worksheet.add_image(excel_image)
            added_keys.add(image.image_key)


def _column_width_pixels(worksheet: Worksheet, column_index: int) -> int:
    """Convert one Excel column width to an approximate pixel width."""

    column_letter = get_column_letter(column_index)
    width = worksheet.column_dimensions[column_letter].width
    if width is None:
        width = worksheet.sheet_format.defaultColWidth or 13
    return max(1, int((width * 7) + 5))


def _row_height_pixels(worksheet: Worksheet, row_index: int) -> int:
    """Convert one Excel row height from points to approximate pixels."""

    height = worksheet.row_dimensions[row_index].height
    if height is None:
        height = worksheet.sheet_format.defaultRowHeight or 15
    return max(1, round(height * 96 / 72))


def _fit_image_to_cell_area(
    worksheet: Worksheet,
    excel_image: ExcelImage,
    anchor_cell: str,
    width_px: int | None,
    height_px: int | None,
) -> None:
    """Scale one image to fit its row and logical module cell area."""

    match = re.fullmatch(r"([A-Za-z]+)(\d+)", anchor_cell)
    if match is None:
        excel_image.anchor = anchor_cell
        return

    column_letters, row_text = match.groups()
    anchor_column = column_index_from_string(column_letters)
    row_index = int(row_text)
    if _WORK_START_COLUMN <= anchor_column <= _WORK_END_COLUMN:
        area_start_column = _WORK_START_COLUMN
        area_end_column = _WORK_END_COLUMN
    else:
        area_start_column = anchor_column
        area_end_column = anchor_column

    available_width = (
        sum(
            _column_width_pixels(worksheet, column_index)
            for column_index in range(area_start_column, area_end_column + 1)
        )
        - 12
    )
    available_height = _row_height_pixels(worksheet, row_index) - 12
    original_width = float(width_px or excel_image.width or 1)
    original_height = float(height_px or excel_image.height or 1)
    scale = min(
        1.0,
        max(1, available_width) / original_width,
        max(1, available_height) / original_height,
    )
    excel_image.width = max(1, round(original_width * scale))
    excel_image.height = max(1, round(original_height * scale))
    excel_image.anchor = anchor_cell


def _set_diff_fill(worksheet: Worksheet, row_index: int, column_index: int) -> None:
    """Apply the requested pale-purple fill to one worksheet cell."""

    worksheet.cell(row=row_index, column=column_index).fill = copy(_DIFF_FILL)


def _fill_work_cells(worksheet: Worksheet, row_index: int, row: ModuleRowData) -> None:
    """Highlight the work cell, including its indentation column."""

    work_column = min(_WORK_START_COLUMN + (row.indent_level or 0), _WORK_END_COLUMN)
    _set_diff_fill(worksheet, row_index, work_column)


def _fill_changed_fields(
    worksheet: Worksheet,
    excel_row: int,
    row: ModuleRowData,
    changed_fields: set[str],
    other_row: ModuleRowData,
    slot_count: int,
) -> None:
    """Highlight only cells whose normalized values differ."""

    scalar_columns = {
        "major_no": 1,
        "middle_no": 2,
        "minor_no": 3,
        "tech_doc_text": 4,
        "expected_result": _EXPECTED_RESULT_COLUMN,
        "note": _EXPECTED_RESULT_COLUMN,
    }
    for field_name, column_index in scalar_columns.items():
        if (
            field_name in changed_fields
            and getattr(row, field_name) != getattr(other_row, field_name)
        ):
            _set_diff_fill(worksheet, excel_row, column_index)

    work_cell_changed = row.work_text != other_row.work_text or (
        row.work_text is not None
        and other_row.work_text is not None
        and row.indent_level != other_row.indent_level
    )
    if work_cell_changed:
        _fill_work_cells(worksheet, excel_row, row)

    if not (
        {"device_entries", "time_text", "window_text", "p_text", "command_text"}
        & changed_fields
    ):
        return

    entries = _entry_map(row)
    other_entries = _entry_map(other_row)
    for slot_no in range(1, slot_count + 1):
        values = entries.get(slot_no, (None, None, None, None))
        other_values = other_entries.get(slot_no, (None, None, None, None))
        start_column = _DEVICE_START_COLUMN + ((slot_no - 1) * _DEVICE_COLUMN_COUNT)
        for offset, (value, other_value) in enumerate(zip(values, other_values, strict=True)):
            if value != other_value:
                _set_diff_fill(worksheet, excel_row, start_column + offset)


def _fill_present_values(
    worksheet: Worksheet,
    excel_row: int,
    row: ModuleRowData,
    slot_count: int,
) -> None:
    """Highlight visible values on an added or removed row, excluding images."""

    scalar_columns = (
        (row.major_no, 1),
        (row.middle_no, 2),
        (row.minor_no, 3),
        (row.tech_doc_text, 4),
        (row.expected_result, _EXPECTED_RESULT_COLUMN),
        (row.note, _EXPECTED_RESULT_COLUMN),
    )
    for value, column_index in scalar_columns:
        if value not in (None, ""):
            _set_diff_fill(worksheet, excel_row, column_index)

    if row.work_text not in (None, ""):
        _fill_work_cells(worksheet, excel_row, row)

    entries = _entry_map(row)
    for slot_no in range(1, slot_count + 1):
        start_column = _DEVICE_START_COLUMN + ((slot_no - 1) * _DEVICE_COLUMN_COUNT)
        for offset, value in enumerate(entries.get(slot_no, (None, None, None, None))):
            if value not in (None, ""):
                _set_diff_fill(worksheet, excel_row, start_column + offset)


def _highlight_diff(
    source_sheet: Worksheet,
    target_sheet: Worksheet,
    source_rows: list[int],
    target_rows: list[int],
    diff: ModuleDiffData,
    slot_count: int,
) -> None:
    """Highlight added, removed, and changed cells on their respective sheets."""

    for diff_row in diff.rows:
        before = diff_row.before
        after = diff_row.after
        if diff_row.status == "added" and after is not None:
            target_row = _excel_row_for_order(target_rows, after.row_order)
            _fill_present_values(target_sheet, target_row, after, slot_count)
            continue

        if diff_row.status == "removed" and before is not None:
            source_row = _excel_row_for_order(source_rows, before.row_order)
            _fill_present_values(source_sheet, source_row, before, slot_count)
            continue

        if diff_row.status != "changed" or before is None or after is None:
            continue

        changed_fields = set(diff_row.changed_fields)
        _fill_changed_fields(
            source_sheet,
            _excel_row_for_order(source_rows, before.row_order),
            before,
            changed_fields,
            after,
            slot_count,
        )
        _fill_changed_fields(
            target_sheet,
            _excel_row_for_order(target_rows, after.row_order),
            after,
            changed_fields,
            before,
            slot_count,
        )


def build_module_diff_workbook(
    diff: ModuleDiffData,
    workbook_bytes: bytes,
    incoming_payload: ModuleCreateRequest,
) -> bytes:
    """Create a two-sheet diff workbook based on the imported workbook layout."""

    if not workbook_bytes:
        raise ValueError("Workbook bytes must not be empty.")

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False, keep_vba=False)
    target_sheet = workbook.worksheets[0]
    target_sheet.title = "_comparison_target"
    source_sheet = workbook.copy_worksheet(target_sheet)
    source_sheet.title = "_comparison_source"

    for worksheet in list(workbook.worksheets):
        if worksheet not in {source_sheet, target_sheet}:
            workbook.remove(worksheet)

    source_sheet.title = "比較元"
    target_sheet.title = "比較先"
    workbook.move_sheet(source_sheet, offset=-1)

    header_row = _find_header_row(target_sheet)
    target_rows = _source_row_indices(target_sheet, incoming_payload, header_row)
    last_column = _tracked_last_column(incoming_payload)
    slot_count = _device_slot_count(incoming_payload)
    reference_row_count = max(
        (row.before.row_order for row in diff.rows if row.before is not None),
        default=0,
    )
    source_rows = _ensure_rows(
        source_sheet,
        list(target_rows),
        reference_row_count,
        last_column,
    )
    _clear_data_cells(source_sheet, source_rows, last_column)

    for diff_row in diff.rows:
        if diff_row.before is None:
            continue
        source_excel_row = _excel_row_for_order(source_rows, diff_row.before.row_order)
        _write_reference_row(
            source_sheet,
            source_excel_row,
            diff_row.before,
            slot_count,
        )

    _add_reference_images(source_sheet, diff)
    _add_incoming_images(target_sheet, incoming_payload)
    _highlight_diff(
        source_sheet,
        target_sheet,
        source_rows,
        target_rows,
        diff,
        slot_count,
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
