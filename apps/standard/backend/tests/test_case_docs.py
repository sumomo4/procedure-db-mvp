"""Tests for case document routes."""

from io import BytesIO
from pathlib import Path

import pytest
from fastapi import status
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.core.config import AppSettings
from app.core.responses import ModuleRowData, SourceDocDetailData, SourceDocModuleItemData


def _tokyo_prefecture(client: TestClient) -> str:
    response = client.get("/api/v1/case-docs/master/prefectures")
    return response.json()["data"]["items"][0]["value"]


def _tokyo_building(client: TestClient) -> str:
    response = client.get(
        "/api/v1/case-docs/master/buildings",
        params={"prefecture": _tokyo_prefecture(client)},
    )
    return response.json()["data"]["items"][0]["value"]


def _fake_source_doc_detail(source_doc_id: int) -> SourceDocDetailData:
    return SourceDocDetailData(
        source_doc_id=source_doc_id,
        source_doc_key="BP-STD-001",
        source_doc_name="M1\u78ba\u8a8d\u7528 \u539f\u672cA",
        description="test source document",
        source_doc_version_id=10,
        version_no=1,
        status="draft",
        status_label="draft",
        change_note=None,
        module_count=1,
        enabled_module_count=1,
        created_by="pytest",
        created_at="2026-05-12",
        updated_at="2026-05-12",
        items=[
            SourceDocModuleItemData(
                blueprint_item_id=100,
                item_order=1,
                enabled=True,
                module_id=1,
                module_key="MOD-001",
                module_name="01.\u30dc\u30fc\u30ec\u30fc\u30c8\u78ba\u8a8d\u30fb\u4fee\u6b63_CS \u30e2\u30b8\u30e5\u30fc\u30eb1",
                module_version_id=101,
                module_version_no=1,
                module_status="draft",
                module_status_label="draft",
                rows=[
                    ModuleRowData(
                        module_row_id=1000,
                        row_order=1,
                        row_type="work",
                        major_no="0",
                        middle_no="1",
                        minor_no="1",
                        tech_doc_text="",
                        work_text="\u4f5c\u696d\u3067\u4f7f\u7528\u3059\u308bPC\u306eTeraTerm\u8a2d\u5b9a\u3092\u5909\u66f4\u3059\u308b\u3002",
                        indent_level=0,
                        expected_result="\u8a2d\u5b9a\u5909\u66f4\u5b8c\u4e86",
                        time_text=None,
                        window_text=None,
                        p_text=None,
                        command_text="{{SBC_COMMAND_FLOATING_IP}}",
                        note=None,
                    ),
                    ModuleRowData(
                        module_row_id=1001,
                        row_order=2,
                        row_type="footer",
                        major_no=None,
                        middle_no=None,
                        minor_no=None,
                        tech_doc_text=None,
                        work_text="\u9023\u7d61\u4e8b\u9805",
                        indent_level=0,
                        expected_result=None,
                        time_text=None,
                        window_text=None,
                        p_text=None,
                        command_text=None,
                        note=None,
                    )
                ],
            )
        ],
    )


def _mock_source_doc_detail(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        "app.routers.case_docs.get_source_doc_detail",
        lambda settings, source_doc_id: _fake_source_doc_detail(source_doc_id),
    )


def _write_placeholder_mapping(path: Path) -> None:
    path.write_text(
        "\n".join(
            [
                "version: 1",
                "placeholders:",
                "- name: LOGIN_USER",
                "  enabled: true",
                "  scope: common",
                "  source_file: case_common_values.xlsx",
                "  key_column: key",
                "  key_value: LOGIN_USER",
                "  value_column: value",
                "  source_column: login_user",
                "  description: login user",
                "",
            ]
        ),
        encoding="utf-8",
    )


def _placeholder_payload(name: str = "TEST_DEVICE_IP") -> dict[str, object]:
    return {
        "name": name,
        "enabled": False,
        "scope": "device",
        "device_type": "SBC",
        "source_file": "SBC.xlsx",
        "key_column": "host_name",
        "value_column": "command_ip",
        "source_column": "command_ip",
        "description": "test placeholder",
    }


def test_read_case_doc_prefectures_returns_options(client: TestClient) -> None:
    """Prefecture master API should return selectable options."""

    response = client.get("/api/v1/case-docs/master/prefectures")

    assert response.status_code == status.HTTP_200_OK
    body = response.json()
    assert body["result"] == "success"
    assert body["data"]["items"]
    assert body["data"]["items"][0]["value"]


def test_read_case_doc_unit_configs_returns_candidates(client: TestClient) -> None:
    """Unit configuration API should filter by selected location."""

    response = client.get(
        "/api/v1/case-docs/master/unit-config",
        params={"prefecture": _tokyo_prefecture(client), "building": _tokyo_building(client)},
    )

    assert response.status_code == status.HTTP_200_OK
    body = response.json()
    assert body["result"] == "success"
    assert body["data"]["items"] == [
        {
            "unit_config_id": "unit-tokyo-001",
            "fs_cluster_name": "FS-CL-TYO-01",
            "block": "B001",
            "prefecture": _tokyo_prefecture(client),
            "building": _tokyo_building(client),
        }
    ]


def test_read_case_doc_placeholder_mappings_returns_configured_items(client: TestClient) -> None:
    """Placeholder mapping API should return configured mapping items."""

    response = client.get("/api/v1/case-docs/placeholders")

    assert response.status_code == status.HTTP_200_OK
    body = response.json()
    assert body["result"] == "success"
    items = body["data"]["items"]
    placeholders = {item["name"]: item for item in items}
    assert placeholders["SBC_COMMAND_FLOATING_IP"]["source_file"] == "SBC.xlsx"
    assert placeholders["SBC_COMMAND_FLOATING_IP"]["scope"] == "device"
    assert placeholders["SBC_COMMAND_FLOATING_IP"]["device_type"] == "SBC"
    assert placeholders["LOGIN_USER"]["scope"] == "common"


def test_validate_case_doc_placeholder_mapping_does_not_write(
    client: TestClient,
    test_settings: AppSettings,
    tmp_path: Path,
) -> None:
    """Validate API should accept a mapping without writing to the YAML file."""

    mapping_path = tmp_path / "placeholder_mapping.yml"
    _write_placeholder_mapping(mapping_path)
    test_settings.case_doc_placeholder_mapping_path = str(mapping_path)

    response = client.post("/api/v1/case-docs/placeholders/validate", json=_placeholder_payload())

    assert response.status_code == status.HTTP_200_OK
    body = response.json()
    assert body["result"] == "success"
    assert body["data"]["name"] == "TEST_DEVICE_IP"
    assert "TEST_DEVICE_IP" not in mapping_path.read_text(encoding="utf-8")


def test_create_case_doc_placeholder_mapping_writes_yaml(
    client: TestClient,
    test_settings: AppSettings,
    tmp_path: Path,
) -> None:
    """Create API should append a placeholder mapping to the YAML file."""

    mapping_path = tmp_path / "placeholder_mapping.yml"
    _write_placeholder_mapping(mapping_path)
    test_settings.case_doc_placeholder_mapping_path = str(mapping_path)

    response = client.post("/api/v1/case-docs/placeholders", json=_placeholder_payload())

    assert response.status_code == status.HTTP_201_CREATED
    body = response.json()
    assert body["result"] == "success"
    assert body["data"]["name"] == "TEST_DEVICE_IP"

    list_response = client.get("/api/v1/case-docs/placeholders")
    placeholders = {item["name"]: item for item in list_response.json()["data"]["items"]}
    assert placeholders["TEST_DEVICE_IP"]["value_column"] == "command_ip"
    assert "TEST_DEVICE_IP" in mapping_path.read_text(encoding="utf-8")


def test_create_case_doc_placeholder_mapping_rejects_duplicate_name(
    client: TestClient,
    test_settings: AppSettings,
    tmp_path: Path,
) -> None:
    """Create API should reject duplicate placeholder names."""

    mapping_path = tmp_path / "placeholder_mapping.yml"
    _write_placeholder_mapping(mapping_path)
    test_settings.case_doc_placeholder_mapping_path = str(mapping_path)

    response = client.post("/api/v1/case-docs/placeholders", json=_placeholder_payload("LOGIN_USER"))

    assert response.status_code == status.HTTP_400_BAD_REQUEST


def test_update_case_doc_placeholder_mapping_writes_yaml(
    client: TestClient,
    test_settings: AppSettings,
    tmp_path: Path,
) -> None:
    """Update API should replace an existing placeholder mapping."""

    mapping_path = tmp_path / "placeholder_mapping.yml"
    _write_placeholder_mapping(mapping_path)
    test_settings.case_doc_placeholder_mapping_path = str(mapping_path)
    client.post("/api/v1/case-docs/placeholders", json=_placeholder_payload())

    payload = _placeholder_payload("TEST_DEVICE_COMMAND_IP")
    payload["description"] = "updated placeholder"
    response = client.put("/api/v1/case-docs/placeholders/TEST_DEVICE_IP", json=payload)

    assert response.status_code == status.HTTP_200_OK
    body = response.json()
    assert body["data"]["name"] == "TEST_DEVICE_COMMAND_IP"
    placeholders = {item["name"]: item for item in client.get("/api/v1/case-docs/placeholders").json()["data"]["items"]}
    assert "TEST_DEVICE_IP" not in placeholders
    assert placeholders["TEST_DEVICE_COMMAND_IP"]["description"] == "updated placeholder"


def test_set_case_doc_placeholder_mapping_enabled_writes_yaml(
    client: TestClient,
    test_settings: AppSettings,
    tmp_path: Path,
) -> None:
    """Enabled API should toggle a placeholder mapping without deleting it."""

    mapping_path = tmp_path / "placeholder_mapping.yml"
    _write_placeholder_mapping(mapping_path)
    test_settings.case_doc_placeholder_mapping_path = str(mapping_path)
    client.post("/api/v1/case-docs/placeholders", json=_placeholder_payload())

    response = client.patch("/api/v1/case-docs/placeholders/TEST_DEVICE_IP/enabled", json={"enabled": True})

    assert response.status_code == status.HTTP_200_OK
    body = response.json()
    assert body["data"]["enabled"] is True
    placeholders = {item["name"]: item for item in client.get("/api/v1/case-docs/placeholders").json()["data"]["items"]}
    assert placeholders["TEST_DEVICE_IP"]["enabled"] is True


def test_resolve_case_doc_context_returns_no_manual_values(client: TestClient) -> None:
    """Resolve context API should return values derived from master data."""

    response = client.post(
        "/api/v1/case-docs/resolve-context",
        json={
            "source_doc_id": 1,
            "prefecture": _tokyo_prefecture(client),
            "building": _tokyo_building(client),
            "unit_config_id": "unit-tokyo-001",
        },
    )

    assert response.status_code == status.HTTP_200_OK
    body = response.json()
    assert body["result"] == "success"
    data = body["data"]
    assert data["unit_config"]["unit_config_id"] == "unit-tokyo-001"
    assert data["target_assignment"]["slot_key"] == "SBC_CL1_0"
    assert data["target_assignment"]["host_name"] == "sbc-tyo-cl1-0"
    assert [item["slot_key"] for item in data["target_assignments"]] == ["SBC_CL1_0"]
    assert data["host_assignments"]
    assert data["common_values"] == [
        {
            "key": "LOGIN_USER",
            "value": "cs-operator",
            "source_table": "case_common_values",
            "source_column": "login_user",
            "source": "case_common_values.login_user",
        }
    ]
    target_placeholders = {
        item["placeholder"]: item
        for item in data["resolved_placeholders"]
        if item["host_name"] == "sbc-tyo-cl1-0"
    }
    assert target_placeholders["SBC_COMMAND_FLOATING_IP"]["value"] == "10.10.1.10"
    assert target_placeholders["SBC_CALL_PROCESS_FLOATING_IP"]["value"] == "10.10.2.10"
    assert target_placeholders["SBC_MAINT_ALARM_LAN_FLOATING_IP"]["value"] == "10.10.3.10"
    assert target_placeholders["SBC_REMOTE_SHELL_FLOATING_IP"]["value"] == "10.10.4.10"
    assert target_placeholders["SBC_NTP_FLOATING_IP"]["value"] == "10.10.5.10"
    assert target_placeholders["TTS_HOST"]["value"] == "tts-tyo-01"
    assert target_placeholders["TTS_IP"]["value"] == "10.10.1.200"
    assert target_placeholders["TTS_PORT"]["value"] == "23"
    login_user_placeholder = next(item for item in data["resolved_placeholders"] if item["placeholder"] == "LOGIN_USER")
    assert login_user_placeholder["source_table"] == "case_common_values"
    assert login_user_placeholder["source_column"] == "login_user"


def test_resolve_case_doc_context_accepts_multiple_target_sbc_slots(client: TestClient) -> None:
    """Resolve context API should accept multiple selected target SBC slots."""

    response = client.post(
        "/api/v1/case-docs/resolve-context",
        json={
            "source_doc_id": 1,
            "prefecture": _tokyo_prefecture(client),
            "building": _tokyo_building(client),
            "unit_config_id": "unit-tokyo-001",
            "target_slot_keys": ["SBC_CL1_1", "SBC_CL1_0"],
        },
    )

    assert response.status_code == status.HTTP_200_OK
    data = response.json()["data"]
    assert data["target_assignment"]["slot_key"] == "SBC_CL1_1"
    assert [item["slot_key"] for item in data["target_assignments"]] == ["SBC_CL1_1", "SBC_CL1_0"]
    assert [item["host_name"] for item in data["target_assignments"]] == ["sbc-tyo-cl1-1", "sbc-tyo-cl1-0"]


def test_resolve_case_doc_context_target_slot_keys_takes_precedence_over_legacy_key(client: TestClient) -> None:
    """Resolve context API should prefer target_slot_keys when both fields are provided."""

    response = client.post(
        "/api/v1/case-docs/resolve-context",
        json={
            "source_doc_id": 1,
            "prefecture": _tokyo_prefecture(client),
            "building": _tokyo_building(client),
            "unit_config_id": "unit-tokyo-001",
            "target_slot_key": "SBC_CL1_0",
            "target_slot_keys": ["SBC_CL1_1"],
        },
    )

    assert response.status_code == status.HTTP_200_OK
    data = response.json()["data"]
    assert data["target_assignment"]["slot_key"] == "SBC_CL1_1"
    assert [item["slot_key"] for item in data["target_assignments"]] == ["SBC_CL1_1"]


def test_resolve_case_doc_context_rejects_unknown_unit(client: TestClient) -> None:
    """Resolve context API should reject unknown selected master values."""

    response = client.post(
        "/api/v1/case-docs/resolve-context",
        json={
            "source_doc_id": 1,
            "prefecture": _tokyo_prefecture(client),
            "building": _tokyo_building(client),
            "unit_config_id": "missing",
        },
    )

    assert response.status_code == status.HTTP_400_BAD_REQUEST
    assert response.json()["result"] == "error"


def test_resolve_case_doc_context_rejects_non_sbc_target_slot(client: TestClient) -> None:
    """Resolve context API should reject a non-SBC target slot."""

    response = client.post(
        "/api/v1/case-docs/resolve-context",
        json={
            "source_doc_id": 1,
            "prefecture": _tokyo_prefecture(client),
            "building": _tokyo_building(client),
            "unit_config_id": "unit-tokyo-001",
            "target_slot_key": "GUI_0",
        },
    )

    assert response.status_code == status.HTTP_400_BAD_REQUEST
    body = response.json()
    assert body["result"] == "error"
    assert body["message"] == "target SBC slot was not found."


def test_generate_case_doc_returns_xlsm_download(client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    """Generate API should return a downloadable workbook package."""

    _mock_source_doc_detail(monkeypatch)

    response = client.post(
        "/api/v1/case-docs/generate",
        json={
            "source_doc_id": 1,
            "prefecture": _tokyo_prefecture(client),
            "building": _tokyo_building(client),
            "unit_config_id": "unit-tokyo-001",
        },
    )

    assert response.status_code == status.HTTP_200_OK
    assert response.headers["content-type"] == "application/vnd.ms-excel.sheet.macroEnabled.12"
    assert response.headers["content-disposition"] == 'attachment; filename="case-doc-1-unit-tokyo-001.xlsm"'

    workbook = load_workbook(BytesIO(response.content), keep_vba=True)
    assert "01.\u30dc\u30fc\u30ec\u30fc\u30c8\u78ba\u8a8d\u30fb\u4fee\u6b63_CS" in workbook.sheetnames
    assert "\u89e3\u6c7a\u5024" in workbook.sheetnames
    template_sheet = workbook["01.\u30dc\u30fc\u30ec\u30fc\u30c8\u78ba\u8a8d\u30fb\u4fee\u6b63_CS"]
    assert template_sheet["M4"].value == "sbc-tyo-cl1-0"
    assert template_sheet["A6"].value == "0"
    assert template_sheet["B6"].value == "1"
    assert template_sheet["C6"].value == "1"
    assert template_sheet["E6"].value == "\u4f5c\u696d\u3067\u4f7f\u7528\u3059\u308bPC\u306eTeraTerm\u8a2d\u5b9a\u3092\u5909\u66f4\u3059\u308b\u3002"
    assert template_sheet["M6"].value == "10.10.1.10"
    assert template_sheet["A6"].border.bottom.style == "thick"
    assert template_sheet["M6"].border.bottom.style == "thick"
    assert template_sheet["A7"].value is None
    assert template_sheet["A8"].value is None
    assert template_sheet["A9"].value is None
    assert template_sheet["A10"].value == "\u9023\u7d61\u4e8b\u9805"
    assert template_sheet["E10"].value is None
    assert all(merged_range.min_row < 6 for merged_range in template_sheet.merged_cells.ranges)
    resolved_sheet = workbook["\u89e3\u6c7a\u5024"]
    assert resolved_sheet["A1"].value == "\u6848\u4ef6CS \u751f\u6210\u7d50\u679c"
    assert resolved_sheet["A22"].value == "SBC_COMMAND_FLOATING_IP"
    target_resolved_values = {
        resolved_sheet.cell(row=row_index, column=1).value: resolved_sheet.cell(row=row_index, column=2).value
        for row_index in range(1, resolved_sheet.max_row + 1)
        if resolved_sheet.cell(row=row_index, column=5).value == "sbc-tyo-cl1-0"
    }
    assert target_resolved_values["SBC_CALL_PROCESS_FLOATING_IP"] == "10.10.2.10"
    assert target_resolved_values["SBC_MAINT_ALARM_LAN_FLOATING_IP"] == "10.10.3.10"
    assert target_resolved_values["SBC_REMOTE_SHELL_FLOATING_IP"] == "10.10.4.10"
    assert target_resolved_values["SBC_NTP_FLOATING_IP"] == "10.10.5.10"
    assert target_resolved_values["TTS_HOST"] == "tts-tyo-01"
    assert target_resolved_values["TTS_IP"] == "10.10.1.200"
    assert target_resolved_values["TTS_PORT"] == "23"
    source_doc_sheet = workbook["\u539f\u672c\u5c55\u958b"]
    assert source_doc_sheet["B2"].value == 1
    assert source_doc_sheet["B3"].value == "BP-STD-001"
    assert source_doc_sheet["I11"].value == "\u4f5c\u696d\u3067\u4f7f\u7528\u3059\u308bPC\u306eTeraTerm\u8a2d\u5b9a\u3092\u5909\u66f4\u3059\u308b\u3002"
    source_doc_values = [
        cell.value
        for row in source_doc_sheet.iter_rows()
        for cell in row
    ]
    assert "10.10.1.10" in source_doc_values
    remaining_placeholders = [
        cell.value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and "{{" in cell.value
    ]
    assert remaining_placeholders == []


def test_resolve_case_doc_context_accepts_target_sbc_slot(client: TestClient) -> None:
    """Resolve context API should use the selected target SBC slot."""

    response = client.post(
        "/api/v1/case-docs/resolve-context",
        json={
            "source_doc_id": 1,
            "prefecture": _tokyo_prefecture(client),
            "building": _tokyo_building(client),
            "unit_config_id": "unit-tokyo-001",
            "target_slot_key": "SBC_CL1_1",
        },
    )

    assert response.status_code == status.HTTP_200_OK
    data = response.json()["data"]
    assert data["target_assignment"]["slot_key"] == "SBC_CL1_1"
    assert data["target_assignment"]["host_name"] == "sbc-tyo-cl1-1"


def test_generate_case_doc_uses_selected_target_sbc_slot(client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    """Generate API should place selected target SBC values into the template."""

    _mock_source_doc_detail(monkeypatch)

    response = client.post(
        "/api/v1/case-docs/generate",
        json={
            "source_doc_id": 1,
            "prefecture": _tokyo_prefecture(client),
            "building": _tokyo_building(client),
            "unit_config_id": "unit-tokyo-001",
            "target_slot_key": "SBC_CL1_1",
        },
    )

    assert response.status_code == status.HTTP_200_OK
    workbook = load_workbook(BytesIO(response.content), keep_vba=True)
    template_sheet = workbook["01.\u30dc\u30fc\u30ec\u30fc\u30c8\u78ba\u8a8d\u30fb\u4fee\u6b63_CS"]
    assert template_sheet["M4"].value == "sbc-tyo-cl1-1"
    assert template_sheet["M6"].value == "10.10.1.11"


def test_generate_case_doc_expands_multiple_target_sbc_blocks(client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    """Generate API should expand J:M target blocks for selected SBC slots."""

    _mock_source_doc_detail(monkeypatch)

    response = client.post(
        "/api/v1/case-docs/generate",
        json={
            "source_doc_id": 1,
            "prefecture": _tokyo_prefecture(client),
            "building": _tokyo_building(client),
            "unit_config_id": "unit-tokyo-001",
            "target_slot_keys": ["SBC_CL1_0", "SBC_CL1_1"],
        },
    )

    assert response.status_code == status.HTTP_200_OK
    workbook = load_workbook(BytesIO(response.content), keep_vba=True)
    template_sheet = workbook["01.\u30dc\u30fc\u30ec\u30fc\u30c8\u78ba\u8a8d\u30fb\u4fee\u6b63_CS"]
    assert template_sheet["M4"].value == "sbc-tyo-cl1-0"
    assert template_sheet["Q4"].value == "sbc-tyo-cl1-1"
    assert template_sheet["J5"].value == "\u6642\u523b"
    assert template_sheet["N5"].value == "\u6642\u523b"
    assert template_sheet["M5"].value == "\u30b3\u30de\u30f3\u30c9"
    assert template_sheet["Q5"].value == "\u30b3\u30de\u30f3\u30c9"
    assert template_sheet["M6"].value == "10.10.1.10"
    assert template_sheet["Q6"].value == "10.10.1.11"
