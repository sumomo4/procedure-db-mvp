"""Case document master repositories.

The seed repository keeps the current deterministic data while the public
case_docs module owns repository selection.
"""

import csv
import re
from collections.abc import Iterable
from pathlib import Path
from typing import Protocol

import yaml
from openpyxl import load_workbook

from app.core.responses import (
    CaseDocCommonValueData,
    CaseDocHostAssignmentData,
    CaseDocMasterOptionData,
    CaseDocPlaceholderMappingEnabledRequest,
    CaseDocPlaceholderMappingItemData,
    CaseDocPlaceholderMappingListData,
    CaseDocPlaceholderMappingUpsertRequest,
    CaseDocMasterOptionsData,
    CaseDocResolvedPlaceholderData,
    CaseDocResolveContextData,
    CaseDocResolveContextRequest,
    CaseDocUnitConfigItemData,
    CaseDocUnitConfigListData,
)


def _u(value: str) -> str:
    """Decode escaped Japanese literals while keeping this source ASCII-safe."""

    return value.encode("ascii").decode("unicode_escape")


_UNIT_CONFIGS = [
    {
        "unit_config_id": "unit-tokyo-001",
        "fs_cluster_name": "FS-CL-TYO-01",
        "block": "B001",
        "prefecture": _u("\\u6771\\u4eac\\u90fd"),
        "building": _u("\\u54c1\\u5ddd\\u30d3\\u30eb"),
        "hosts": {
            "GUI_0": {"device_type": "GUI", "system": "0", "host_name": "gui-tyo-001-0"},
            "GUI_1": {"device_type": "GUI", "system": "1", "host_name": "gui-tyo-001-1"},
            "SBC_CL1_0": {"device_type": "SBC", "system": "0", "host_name": "sbc-tyo-cl1-0"},
            "SBC_CL1_1": {"device_type": "SBC", "system": "1", "host_name": "sbc-tyo-cl1-1"},
        },
    },
    {
        "unit_config_id": "unit-osaka-001",
        "fs_cluster_name": "FS-CL-OSA-01",
        "block": "B002",
        "prefecture": _u("\\u5927\\u962a\\u5e9c"),
        "building": _u("\\u5802\\u5cf6\\u30d3\\u30eb"),
        "hosts": {
            "GUI_0": {"device_type": "GUI", "system": "0", "host_name": "gui-osa-001-0"},
            "GUI_1": {"device_type": "GUI", "system": "1", "host_name": "gui-osa-001-1"},
            "SBC_CL1_0": {"device_type": "SBC", "system": "0", "host_name": "sbc-osa-cl1-0"},
            "SBC_CL1_1": {"device_type": "SBC", "system": "1", "host_name": "sbc-osa-cl1-1"},
        },
    },
]

_DEVICE_VALUES_BY_HOST_NAME = {
    "sbc-tyo-cl1-0": {
        "command_floating_ip": "10.10.1.10",
        "call_process_floating_ip": "10.10.2.10",
        "maint_alarm_lan_floating_ip": "10.10.3.10",
        "remote_shell_floating_ip": "10.10.4.10",
        "ntp_floating_ip": "10.10.5.10",
        "tts_host": "tts-tyo-01",
        "tts_ip": "10.10.1.200",
        "tts_port": "23",
    },
    "sbc-tyo-cl1-1": {
        "command_floating_ip": "10.10.1.11",
        "call_process_floating_ip": "10.10.2.11",
        "maint_alarm_lan_floating_ip": "10.10.3.11",
        "remote_shell_floating_ip": "10.10.4.11",
        "ntp_floating_ip": "10.10.5.11",
        "tts_host": "tts-tyo-01",
        "tts_ip": "10.10.1.200",
        "tts_port": "23",
    },
    "sbc-osa-cl1-0": {
        "command_floating_ip": "10.20.1.10",
        "call_process_floating_ip": "10.20.2.10",
        "maint_alarm_lan_floating_ip": "10.20.3.10",
        "remote_shell_floating_ip": "10.20.4.10",
        "ntp_floating_ip": "10.20.5.10",
        "tts_host": "tts-osa-01",
        "tts_ip": "10.20.1.200",
        "tts_port": "23",
    },
    "sbc-osa-cl1-1": {
        "command_floating_ip": "10.20.1.11",
        "call_process_floating_ip": "10.20.2.11",
        "maint_alarm_lan_floating_ip": "10.20.3.11",
        "remote_shell_floating_ip": "10.20.4.11",
        "ntp_floating_ip": "10.20.5.11",
        "tts_host": "tts-osa-01",
        "tts_ip": "10.20.1.200",
        "tts_port": "23",
    },
}

_COMMON_VALUE_SOURCE_TABLE = "case_common_values"
_COMMON_VALUES_BY_KEY = {"LOGIN_USER": "cs-operator"}

UNIT_CONFIG_FILE_NAMES = ("unit_config.xlsx", _u(r"\u30e6\u30cb\u30c3\u30c8\u69cb\u6210.xlsx"))
SBC_FILE_NAMES = ("SBC.xlsx", "sbc.xlsx")
COMMON_VALUES_XLSX_FILE_NAME = "case_common_values.xlsx"
COMMON_VALUES_CSV_FILE_NAME = "case_common_values.csv"

UNIT_CONFIG_COLUMN_ALIASES = {
    "unit_config_id": ("unit_config_id", _u(r"\u30e6\u30cb\u30c3\u30c8\u69cb\u6210ID")),
    "fs_cluster_name": ("fs_cluster_name", "FS" + _u(r"\u30af\u30e9\u30b9\u30bf\u540d")),
    "block": ("block", _u(r"\u30d6\u30ed\u30c3\u30af")),
    "prefecture": (
        "prefecture",
        _u(r"\u90fd\u9053\u5e9c\u770c"),
        _u(r"\u88c5\u7f6e\u8a2d\u7f6e\u90fd\u9053\u5e9c\u770c"),
        _u(r"\u88c5\u7f6e\u8a2d\u7f6e\u5e9c\u770c"),
    ),
    "building": ("building", _u(r"\u30d3\u30eb"), _u(r"\u88c5\u7f6e\u8a2d\u7f6e\u30d3\u30eb")),
}
SBC_COLUMN_ALIASES = {
    "host_name": ("host_name", _u(r"\u30db\u30b9\u30c8\u540d")),
    "command_floating_ip": (
        "command_floating_ip",
        _u(r"\u30b3\u30de\u30f3\u30c9\u7528\u30d5\u30ed\u30fc\u30c6\u30a3\u30f3\u30b0IP\u30a2\u30c9\u30ec\u30b9"),
    ),
    "call_process_floating_ip": (
        "call_process_floating_ip",
        _u(r"\u547c\u51e6\u7406\u7528\u30d5\u30ed\u30fc\u30c6\u30a3\u30f3\u30b0IP\u30a2\u30c9\u30ec\u30b9"),
    ),
    "maint_alarm_lan_floating_ip": (
        "maint_alarm_lan_floating_ip",
        _u(r"\u4fdd\u5b88\u30a2\u30e9\u30fc\u30e0\u7528LAN\u30d5\u30ed\u30fc\u30c6\u30a3\u30f3\u30b0IP\u30a2\u30c9\u30ec\u30b9"),
    ),
    "remote_shell_floating_ip": (
        "remote_shell_floating_ip",
        _u(r"\u30ea\u30e2\u30fc\u30c8\u30b7\u30a7\u30eb\u30b3\u30de\u30f3\u30c9\u7528\u30d5\u30ed\u30fc\u30c6\u30a3\u30f3\u30b0IP\u30a2\u30c9\u30ec\u30b9"),
    ),
    "ntp_floating_ip": (
        "ntp_floating_ip",
        _u(r"NTP\u5411\u3051\u30d5\u30ed\u30fc\u30c6\u30a3\u30f3\u30b0IP\u30a2\u30c9\u30ec\u30b9"),
    ),
    "tts_host": ("tts_host", "TTS-Host"),
    "tts_ip": ("tts_ip", "TTS-IP"),
    "tts_port": ("tts_port", "TTS-Port"),
}
COMMON_VALUE_COLUMN_ALIASES = {
    "key": ("key", _u(r"\u30ad\u30fc")),
    "value": ("value", _u(r"\u5024")),
    "source_table": ("source_table", _u(r"\u51fa\u5178\u30c6\u30fc\u30d6\u30eb")),
    "source_column": ("source_column", _u(r"\u51fa\u5178\u30ab\u30e9\u30e0")),
}
DEVICE_SLOT_SUFFIX = _u(r"\u7cfb")
PLACEHOLDER_NAME_PATTERN = re.compile(r"^[A-Z0-9_]+$")


def _resolve_placeholder_mapping_path(mapping_path: str) -> Path:
    path = Path(mapping_path)
    if not path.is_absolute():
        path = Path.cwd() / path
    return path


def _load_placeholder_mapping_payload(mapping_path: str) -> dict[str, object]:
    path = _resolve_placeholder_mapping_path(mapping_path)
    if not path.exists():
        raise ValueError(f"placeholder mapping file was not found: {path}")

    with path.open(encoding="utf-8-sig") as file:
        payload = yaml.safe_load(file) or {}
    if not isinstance(payload, dict):
        raise ValueError("placeholder mapping file must contain an object.")
    return payload


def _validate_placeholder_mapping_item(item: CaseDocPlaceholderMappingItemData) -> None:
    if not PLACEHOLDER_NAME_PATTERN.fullmatch(item.name):
        raise ValueError(f"placeholder mapping name is invalid: {item.name}")
    if item.scope == "device" and not item.device_type:
        raise ValueError(f"device placeholder requires device_type: {item.name}")
    if item.scope == "common" and not item.key_value:
        raise ValueError(f"common placeholder requires key_value: {item.name}")


def _validate_placeholder_mappings(mappings: Iterable[CaseDocPlaceholderMappingItemData]) -> list[CaseDocPlaceholderMappingItemData]:
    validated = list(mappings)
    seen: set[str] = set()
    for item in validated:
        _validate_placeholder_mapping_item(item)
        if item.name in seen:
            raise ValueError(f"placeholder mapping name is duplicated: {item.name}")
        seen.add(item.name)
    return validated


def _load_placeholder_mappings(mapping_path: str) -> list[CaseDocPlaceholderMappingItemData]:
    payload = _load_placeholder_mapping_payload(mapping_path)

    raw_items = payload.get("placeholders", [])
    if not isinstance(raw_items, list):
        raise ValueError("placeholder mapping file must contain a placeholders list.")

    mappings: list[CaseDocPlaceholderMappingItemData] = []
    for raw_item in raw_items:
        if not isinstance(raw_item, dict):
            raise ValueError("placeholder mapping entries must be objects.")
        mappings.append(CaseDocPlaceholderMappingItemData(**raw_item))
    return _validate_placeholder_mappings(mappings)


def _placeholder_mapping_to_dict(item: CaseDocPlaceholderMappingItemData) -> dict[str, object]:
    payload: dict[str, object] = {
        "name": item.name,
        "enabled": item.enabled,
        "scope": item.scope,
    }
    if item.device_type is not None:
        payload["device_type"] = item.device_type
    payload.update(
        {
            "source_file": item.source_file,
            "key_column": item.key_column,
        }
    )
    if item.key_value is not None:
        payload["key_value"] = item.key_value
    payload.update(
        {
            "value_column": item.value_column,
            "source_column": item.source_column,
        }
    )
    if item.description is not None:
        payload["description"] = item.description
    return payload


def _write_placeholder_mappings(
    mapping_path: str,
    mappings: Iterable[CaseDocPlaceholderMappingItemData],
) -> CaseDocPlaceholderMappingListData:
    path = _resolve_placeholder_mapping_path(mapping_path)
    validated = _validate_placeholder_mappings(mappings)
    payload = {
        "version": 1,
        "placeholders": [_placeholder_mapping_to_dict(item) for item in validated],
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as file:
        yaml.safe_dump(payload, file, allow_unicode=True, sort_keys=False)
    return CaseDocPlaceholderMappingListData(items=validated)


def _to_placeholder_mapping_item(
    payload: CaseDocPlaceholderMappingUpsertRequest,
) -> CaseDocPlaceholderMappingItemData:
    return CaseDocPlaceholderMappingItemData(**payload.model_dump())


def _create_placeholder_mapping(
    mapping_path: str,
    payload: CaseDocPlaceholderMappingUpsertRequest,
) -> CaseDocPlaceholderMappingItemData:
    mappings = _load_placeholder_mappings(mapping_path)
    item = _to_placeholder_mapping_item(payload)
    if any(existing.name == item.name for existing in mappings):
        raise ValueError(f"placeholder mapping name is duplicated: {item.name}")
    _validate_placeholder_mapping_item(item)
    mappings.append(item)
    _write_placeholder_mappings(mapping_path, mappings)
    return item


def _update_placeholder_mapping(
    mapping_path: str,
    name: str,
    payload: CaseDocPlaceholderMappingUpsertRequest,
) -> CaseDocPlaceholderMappingItemData:
    mappings = _load_placeholder_mappings(mapping_path)
    item = _to_placeholder_mapping_item(payload)
    _validate_placeholder_mapping_item(item)
    matched = False
    updated: list[CaseDocPlaceholderMappingItemData] = []
    for existing in mappings:
        if existing.name == name:
            updated.append(item)
            matched = True
        else:
            updated.append(existing)
    if not matched:
        raise ValueError(f"placeholder mapping was not found: {name}")
    _write_placeholder_mappings(mapping_path, updated)
    return item


def _set_placeholder_mapping_enabled(
    mapping_path: str,
    name: str,
    payload: CaseDocPlaceholderMappingEnabledRequest,
) -> CaseDocPlaceholderMappingItemData:
    mappings = _load_placeholder_mappings(mapping_path)
    matched = False
    updated_item: CaseDocPlaceholderMappingItemData | None = None
    updated: list[CaseDocPlaceholderMappingItemData] = []
    for existing in mappings:
        if existing.name == name:
            updated_item = existing.model_copy(update={"enabled": payload.enabled})
            updated.append(updated_item)
            matched = True
        else:
            updated.append(existing)
    if not matched or updated_item is None:
        raise ValueError(f"placeholder mapping was not found: {name}")
    _write_placeholder_mappings(mapping_path, updated)
    return updated_item


def _enabled_device_mappings(
    mappings: Iterable[CaseDocPlaceholderMappingItemData],
    device_type: str,
) -> list[CaseDocPlaceholderMappingItemData]:
    return [
        mapping
        for mapping in mappings
        if mapping.enabled and mapping.scope == "device" and mapping.device_type == device_type
    ]


def _enabled_common_mappings(
    mappings: Iterable[CaseDocPlaceholderMappingItemData],
) -> list[CaseDocPlaceholderMappingItemData]:
    return [mapping for mapping in mappings if mapping.enabled and mapping.scope == "common"]


def _source_table_from_mapping(mapping: CaseDocPlaceholderMappingItemData) -> str:
    return Path(mapping.source_file).stem


def _normalize_key(value: str) -> str:
    return re.sub(r"[\s\u3000]+", "", value).lower()


def _resolve_export_file_path(export_dir: Path, file_names: Iterable[str]) -> Path:
    for file_name in file_names:
        path = export_dir / file_name
        if path.exists():
            return path
    expected = ", ".join(file_names)
    raise ValueError(f"case document export file was not found: {expected}")


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _read_xlsx_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise ValueError(f"case document export file was not found: {path.name}")

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_index = next((index for index, row in enumerate(rows) if any(_cell_to_text(value) for value in row)), None)
    if header_index is None:
        return []

    headers = [_normalize_key(_cell_to_text(value)) for value in rows[header_index]]
    records: list[dict[str, str]] = []
    for row in rows[header_index + 1 :]:
        record = {header: _cell_to_text(value) for header, value in zip(headers, row, strict=False) if header}
        if any(record.values()):
            records.append(record)
    return records


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise ValueError(f"case document export file was not found: {path.name}")

    with path.open(encoding="utf-8-sig", newline="") as file:
        return [
            {_normalize_key(str(key)): _cell_to_text(value) for key, value in row.items() if key}
            for row in csv.DictReader(file)
            if any(_cell_to_text(value) for value in row.values())
        ]


def _value_from_aliases(row: dict[str, str], aliases: Iterable[str], field_name: str) -> str:
    for alias in aliases:
        value = row.get(_normalize_key(alias), "").strip()
        if value:
            return value
    raise ValueError(f"required column value was not found: {field_name}")


def _optional_value_from_aliases(row: dict[str, str], aliases: Iterable[str]) -> str:
    for alias in aliases:
        value = row.get(_normalize_key(alias), "").strip()
        if value:
            return value
    return ""


def _slot_key_from_column_name(column_name: str) -> str | None:
    ignored_columns = {_normalize_key(alias) for aliases in UNIT_CONFIG_COLUMN_ALIASES.values() for alias in aliases}
    if column_name in ignored_columns:
        return None
    normalized_suffix = _normalize_key(DEVICE_SLOT_SUFFIX)
    if not column_name.endswith(normalized_suffix):
        return None
    slot_key = column_name.removesuffix(normalized_suffix).strip().upper()
    return slot_key or None


def _as_option(value: str) -> CaseDocMasterOptionData:
    return CaseDocMasterOptionData(value=value, label=value)


def _unique_ordered(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        normalized = value.strip()
        if normalized and normalized not in seen:
            seen.add(normalized)
            ordered.append(normalized)
    return ordered


def _sort_unit_configs(unit_configs: Iterable[dict[str, object]]) -> list[dict[str, object]]:
    return sorted(
        unit_configs,
        key=lambda row: (
            str(row["prefecture"]),
            str(row["building"]),
            str(row["fs_cluster_name"]),
            str(row["block"]),
            str(row["unit_config_id"]),
        ),
    )


def _to_unit_config_item(row: dict[str, object]) -> CaseDocUnitConfigItemData:
    return CaseDocUnitConfigItemData(
        unit_config_id=str(row["unit_config_id"]),
        fs_cluster_name=str(row["fs_cluster_name"]),
        block=str(row["block"]),
        prefecture=str(row["prefecture"]),
        building=str(row["building"]),
    )


def _find_unit_config(payload: CaseDocResolveContextRequest) -> dict[str, object]:
    candidates = [
        row
        for row in _UNIT_CONFIGS
        if row["prefecture"] == payload.prefecture and row["building"] == payload.building
    ]

    if payload.unit_config_id:
        candidates = [row for row in candidates if row["unit_config_id"] == payload.unit_config_id]
    if payload.fs_cluster_name:
        candidates = [row for row in candidates if row["fs_cluster_name"] == payload.fs_cluster_name]
    if payload.block:
        candidates = [row for row in candidates if row["block"] == payload.block]

    if not candidates:
        raise ValueError("unit configuration was not found.")
    return candidates[0]


def _select_target_host_assignment(
    host_assignments: list[CaseDocHostAssignmentData],
    target_slot_key: str | None,
) -> CaseDocHostAssignmentData:
    """Select the target SBC host assignment used as the document value key."""

    sbc_assignments = [assignment for assignment in host_assignments if assignment.device_type == "SBC"]
    if target_slot_key:
        for assignment in sbc_assignments:
            if assignment.slot_key == target_slot_key:
                return assignment
        raise ValueError("target SBC slot was not found.")

    if not sbc_assignments:
        raise ValueError("target SBC slot was not found.")
    return sbc_assignments[0]


def _to_host_assignments(hosts: dict[str, object]) -> list[CaseDocHostAssignmentData]:
    return [
        CaseDocHostAssignmentData(
            slot_key=slot_key,
            device_type=str(host["device_type"]),
            system=str(host["system"]),
            host_name=str(host["host_name"]),
        )
        for slot_key, host in hosts.items()
        if isinstance(host, dict)
    ]


def _list_common_values(
    mappings: list[CaseDocPlaceholderMappingItemData],
) -> list[CaseDocCommonValueData]:
    values: list[CaseDocCommonValueData] = []
    for mapping in _enabled_common_mappings(mappings):
        key = mapping.key_value or mapping.name
        value = _COMMON_VALUES_BY_KEY.get(key)
        if value is None:
            continue
        source_table = _source_table_from_mapping(mapping)
        values.append(
            CaseDocCommonValueData(
                key=key,
                value=value,
                source_table=source_table,
                source_column=mapping.source_column,
                source=f"{source_table}.{mapping.source_column}",
            )
        )
    return values


def _resolve_sbc_placeholders_by_host_name(
    host_assignments: list[CaseDocHostAssignmentData],
    mappings: list[CaseDocPlaceholderMappingItemData],
) -> list[CaseDocResolvedPlaceholderData]:
    """Resolve SBC values by using each assignment host name as the master key."""

    return _resolve_device_placeholders_from_values(host_assignments, _DEVICE_VALUES_BY_HOST_NAME, mappings, "SBC")


def _resolve_device_placeholders_from_values(
    host_assignments: list[CaseDocHostAssignmentData],
    device_values_by_host_name: dict[str, dict[str, str]],
    mappings: list[CaseDocPlaceholderMappingItemData],
    device_type: str,
) -> list[CaseDocResolvedPlaceholderData]:
    resolved_placeholders: list[CaseDocResolvedPlaceholderData] = []
    device_mappings = _enabled_device_mappings(mappings, device_type)
    for assignment in host_assignments:
        if assignment.device_type != device_type:
            continue

        device_values = device_values_by_host_name.get(assignment.host_name)
        if device_values is None:
            continue

        for mapping in device_mappings:
            value = device_values.get(mapping.source_column)
            if not value:
                continue
            resolved_placeholders.append(
                CaseDocResolvedPlaceholderData(
                    placeholder=mapping.name,
                    value=value,
                    source_table=_source_table_from_mapping(mapping),
                    source_column=mapping.source_column,
                    host_name=assignment.host_name,
                )
            )
    return resolved_placeholders


def _resolve_common_placeholders(
    common_values: list[CaseDocCommonValueData],
) -> list[CaseDocResolvedPlaceholderData]:
    return [
        CaseDocResolvedPlaceholderData(
            placeholder=value.key,
            value=value.value,
            source_table=value.source_table,
            source_column=value.source_column,
            host_name=None,
        )
        for value in common_values
    ]


class CaseDocMasterRepository(Protocol):
    """Repository boundary for case document master data."""

    def list_prefectures(self) -> CaseDocMasterOptionsData:
        """Return prefectures available for case document generation."""

    def list_buildings(self, prefecture: str) -> CaseDocMasterOptionsData:
        """Return buildings filtered by prefecture."""

    def list_unit_configs(self, prefecture: str, building: str) -> CaseDocUnitConfigListData:
        """Return unit configuration candidates for selected location."""

    def resolve_context(self, payload: CaseDocResolveContextRequest) -> CaseDocResolveContextData:
        """Resolve generation context from repository data."""

    def list_placeholder_mappings(self) -> CaseDocPlaceholderMappingListData:
        """Return placeholder mappings used for case document generation."""

    def validate_placeholder_mapping(
        self,
        payload: CaseDocPlaceholderMappingUpsertRequest,
    ) -> CaseDocPlaceholderMappingItemData:
        """Validate a placeholder mapping without writing it."""

    def create_placeholder_mapping(
        self,
        payload: CaseDocPlaceholderMappingUpsertRequest,
    ) -> CaseDocPlaceholderMappingItemData:
        """Create a placeholder mapping."""

    def update_placeholder_mapping(
        self,
        name: str,
        payload: CaseDocPlaceholderMappingUpsertRequest,
    ) -> CaseDocPlaceholderMappingItemData:
        """Update a placeholder mapping."""

    def set_placeholder_mapping_enabled(
        self,
        name: str,
        payload: CaseDocPlaceholderMappingEnabledRequest,
    ) -> CaseDocPlaceholderMappingItemData:
        """Enable or disable a placeholder mapping."""


class SeedCaseDocMasterRepository:
    """Deterministic case document master data used before Access export import."""

    def __init__(self, placeholder_mapping_path: str = "app/config/placeholder_mapping.yml") -> None:
        self.placeholder_mapping_path = placeholder_mapping_path

    def list_placeholder_mappings(self) -> CaseDocPlaceholderMappingListData:
        return CaseDocPlaceholderMappingListData(items=_load_placeholder_mappings(self.placeholder_mapping_path))

    def validate_placeholder_mapping(
        self,
        payload: CaseDocPlaceholderMappingUpsertRequest,
    ) -> CaseDocPlaceholderMappingItemData:
        item = _to_placeholder_mapping_item(payload)
        _validate_placeholder_mapping_item(item)
        return item

    def create_placeholder_mapping(
        self,
        payload: CaseDocPlaceholderMappingUpsertRequest,
    ) -> CaseDocPlaceholderMappingItemData:
        return _create_placeholder_mapping(self.placeholder_mapping_path, payload)

    def update_placeholder_mapping(
        self,
        name: str,
        payload: CaseDocPlaceholderMappingUpsertRequest,
    ) -> CaseDocPlaceholderMappingItemData:
        return _update_placeholder_mapping(self.placeholder_mapping_path, name, payload)

    def set_placeholder_mapping_enabled(
        self,
        name: str,
        payload: CaseDocPlaceholderMappingEnabledRequest,
    ) -> CaseDocPlaceholderMappingItemData:
        return _set_placeholder_mapping_enabled(self.placeholder_mapping_path, name, payload)

    def _load_placeholder_mappings(self) -> list[CaseDocPlaceholderMappingItemData]:
        return _load_placeholder_mappings(self.placeholder_mapping_path)

    def list_prefectures(self) -> CaseDocMasterOptionsData:
        values = _unique_ordered(str(row["prefecture"]) for row in _UNIT_CONFIGS)
        return CaseDocMasterOptionsData(items=[_as_option(value) for value in values])

    def list_buildings(self, prefecture: str) -> CaseDocMasterOptionsData:
        values = _unique_ordered(
            str(row["building"])
            for row in _UNIT_CONFIGS
            if row["prefecture"] == prefecture
        )
        return CaseDocMasterOptionsData(items=[_as_option(value) for value in values])

    def list_unit_configs(self, prefecture: str, building: str) -> CaseDocUnitConfigListData:
        items = [
            _to_unit_config_item(row)
            for row in _UNIT_CONFIGS
            if row["prefecture"] == prefecture and row["building"] == building
        ]
        return CaseDocUnitConfigListData(items=items)

    def resolve_context(self, payload: CaseDocResolveContextRequest) -> CaseDocResolveContextData:
        unit_config = _find_unit_config(payload)
        hosts = unit_config["hosts"]
        if not isinstance(hosts, dict):
            raise ValueError("unit configuration hosts are invalid.")

        host_assignments = _to_host_assignments(hosts)
        target_assignment = _select_target_host_assignment(host_assignments, payload.target_slot_key)
        mappings = self._load_placeholder_mappings()
        common_values = _list_common_values(mappings)
        resolved_placeholders = [
            *_resolve_sbc_placeholders_by_host_name(host_assignments, mappings),
            *_resolve_common_placeholders(common_values),
        ]

        return CaseDocResolveContextData(
            source_doc_id=payload.source_doc_id,
            unit_config=_to_unit_config_item(unit_config),
            target_assignment=target_assignment,
            host_assignments=host_assignments,
            common_values=common_values,
            resolved_placeholders=resolved_placeholders,
        )


class ExportFileCaseDocMasterRepository:
    """Case document master data loaded from Access-derived export files."""

    def __init__(self, export_dir: str, placeholder_mapping_path: str = "app/config/placeholder_mapping.yml") -> None:
        self.export_dir = Path(export_dir)
        self.placeholder_mapping_path = placeholder_mapping_path

    def list_placeholder_mappings(self) -> CaseDocPlaceholderMappingListData:
        return CaseDocPlaceholderMappingListData(items=_load_placeholder_mappings(self.placeholder_mapping_path))

    def validate_placeholder_mapping(
        self,
        payload: CaseDocPlaceholderMappingUpsertRequest,
    ) -> CaseDocPlaceholderMappingItemData:
        item = _to_placeholder_mapping_item(payload)
        _validate_placeholder_mapping_item(item)
        return item

    def create_placeholder_mapping(
        self,
        payload: CaseDocPlaceholderMappingUpsertRequest,
    ) -> CaseDocPlaceholderMappingItemData:
        return _create_placeholder_mapping(self.placeholder_mapping_path, payload)

    def update_placeholder_mapping(
        self,
        name: str,
        payload: CaseDocPlaceholderMappingUpsertRequest,
    ) -> CaseDocPlaceholderMappingItemData:
        return _update_placeholder_mapping(self.placeholder_mapping_path, name, payload)

    def set_placeholder_mapping_enabled(
        self,
        name: str,
        payload: CaseDocPlaceholderMappingEnabledRequest,
    ) -> CaseDocPlaceholderMappingItemData:
        return _set_placeholder_mapping_enabled(self.placeholder_mapping_path, name, payload)

    def _load_placeholder_mappings(self) -> list[CaseDocPlaceholderMappingItemData]:
        return _load_placeholder_mappings(self.placeholder_mapping_path)

    def list_prefectures(self) -> CaseDocMasterOptionsData:
        values = sorted(_unique_ordered(str(row["prefecture"]) for row in self._load_unit_configs()))
        return CaseDocMasterOptionsData(items=[_as_option(value) for value in values])

    def list_buildings(self, prefecture: str) -> CaseDocMasterOptionsData:
        values = sorted(
            _unique_ordered(
                str(row["building"])
                for row in self._load_unit_configs()
                if row["prefecture"] == prefecture
            )
        )
        return CaseDocMasterOptionsData(items=[_as_option(value) for value in values])

    def list_unit_configs(self, prefecture: str, building: str) -> CaseDocUnitConfigListData:
        items = [
            _to_unit_config_item(row)
            for row in _sort_unit_configs(self._load_unit_configs())
            if row["prefecture"] == prefecture and row["building"] == building
        ]
        return CaseDocUnitConfigListData(items=items)

    def resolve_context(self, payload: CaseDocResolveContextRequest) -> CaseDocResolveContextData:
        unit_config = self._find_unit_config(payload)
        hosts = unit_config["hosts"]
        if not isinstance(hosts, dict):
            raise ValueError("unit configuration hosts are invalid.")

        host_assignments = _to_host_assignments(hosts)
        target_assignment = _select_target_host_assignment(host_assignments, payload.target_slot_key)
        mappings = self._load_placeholder_mappings()
        common_values = self._load_common_values(mappings)
        resolved_placeholders = [
            *self._resolve_device_placeholders_by_host_name(host_assignments, mappings),
            *_resolve_common_placeholders(common_values),
        ]

        return CaseDocResolveContextData(
            source_doc_id=payload.source_doc_id,
            unit_config=_to_unit_config_item(unit_config),
            target_assignment=target_assignment,
            host_assignments=host_assignments,
            common_values=common_values,
            resolved_placeholders=resolved_placeholders,
        )

    def _find_unit_config(self, payload: CaseDocResolveContextRequest) -> dict[str, object]:
        candidates = [
            row
            for row in self._load_unit_configs()
            if row["prefecture"] == payload.prefecture and row["building"] == payload.building
        ]

        if payload.unit_config_id:
            candidates = [row for row in candidates if row["unit_config_id"] == payload.unit_config_id]
        if payload.fs_cluster_name:
            candidates = [row for row in candidates if row["fs_cluster_name"] == payload.fs_cluster_name]
        if payload.block:
            candidates = [row for row in candidates if row["block"] == payload.block]

        if not candidates:
            raise ValueError("unit configuration was not found.")
        return candidates[0]

    def _load_unit_configs(self) -> list[dict[str, object]]:
        rows = _read_xlsx_rows(_resolve_export_file_path(self.export_dir, UNIT_CONFIG_FILE_NAMES))
        unit_configs: list[dict[str, object]] = []
        for index, row in enumerate(rows, start=1):
            fs_cluster_name = _value_from_aliases(row, UNIT_CONFIG_COLUMN_ALIASES["fs_cluster_name"], "fs_cluster_name")
            block = _value_from_aliases(row, UNIT_CONFIG_COLUMN_ALIASES["block"], "block")
            unit_config_id = _optional_value_from_aliases(row, UNIT_CONFIG_COLUMN_ALIASES["unit_config_id"])
            if not unit_config_id:
                unit_config_id = f"unit-export-{index}"

            hosts = {
                slot_key: {
                    "device_type": slot_key.split("_", 1)[0],
                    "system": slot_key.rsplit("_", 1)[-1] if "_" in slot_key else None,
                    "host_name": host_name,
                }
                for column_name, host_name in row.items()
                if (slot_key := _slot_key_from_column_name(column_name)) is not None
                if host_name
            }

            unit_configs.append(
                {
                    "unit_config_id": unit_config_id,
                    "fs_cluster_name": fs_cluster_name,
                    "block": block,
                    "prefecture": _value_from_aliases(row, UNIT_CONFIG_COLUMN_ALIASES["prefecture"], "prefecture"),
                    "building": _value_from_aliases(row, UNIT_CONFIG_COLUMN_ALIASES["building"], "building"),
                    "hosts": hosts,
                }
            )
        return unit_configs

    def _load_device_values_by_host_name(
        self,
        mappings: list[CaseDocPlaceholderMappingItemData],
        device_type: str,
    ) -> dict[str, dict[str, str]]:
        values_by_host_name: dict[str, dict[str, str]] = {}
        device_mappings = _enabled_device_mappings(mappings, device_type)
        source_files = _unique_ordered(mapping.source_file for mapping in device_mappings)
        for source_file in source_files:
            source_mappings = [mapping for mapping in device_mappings if mapping.source_file == source_file]
            rows = _read_xlsx_rows(self.export_dir / source_file)
            for row in rows:
                host_aliases = ["host_name", _u(r"\u30db\u30b9\u30c8\u540d")]
                host_aliases.extend(mapping.key_column for mapping in source_mappings)
                host_name = _value_from_aliases(row, host_aliases, "host_name")
                device_values = values_by_host_name.setdefault(host_name, {})
                for mapping in source_mappings:
                    value = _optional_value_from_aliases(row, (mapping.value_column, mapping.source_column))
                    if value:
                        device_values[mapping.source_column] = value
        return values_by_host_name

    def _load_common_values(
        self,
        mappings: list[CaseDocPlaceholderMappingItemData],
    ) -> list[CaseDocCommonValueData]:
        xlsx_path = self.export_dir / COMMON_VALUES_XLSX_FILE_NAME
        csv_path = self.export_dir / COMMON_VALUES_CSV_FILE_NAME
        if xlsx_path.exists():
            rows = _read_xlsx_rows(xlsx_path)
        elif csv_path.exists():
            rows = _read_csv_rows(csv_path)
        else:
            return _list_common_values(mappings)

        common_values: list[CaseDocCommonValueData] = []
        for mapping in _enabled_common_mappings(mappings):
            matched_row = next(
                (
                    row
                    for row in rows
                    if _optional_value_from_aliases(row, (mapping.key_column, "key")) == mapping.key_value
                ),
                None,
            )
            if matched_row is None:
                fallback_value = _COMMON_VALUES_BY_KEY.get(mapping.key_value or mapping.name)
                if fallback_value is None:
                    continue
                source_table = _source_table_from_mapping(mapping)
                common_values.append(
                    CaseDocCommonValueData(
                        key=mapping.key_value or mapping.name,
                        value=fallback_value,
                        source_table=source_table,
                        source_column=mapping.source_column,
                        source=f"{source_table}.{mapping.source_column}",
                    )
                )
                continue

            key = mapping.key_value or mapping.name
            value = _value_from_aliases(matched_row, (mapping.value_column, "value"), mapping.name)
            source_table = _optional_value_from_aliases(matched_row, COMMON_VALUE_COLUMN_ALIASES["source_table"]) or _source_table_from_mapping(mapping)
            source_column = _optional_value_from_aliases(matched_row, COMMON_VALUE_COLUMN_ALIASES["source_column"]) or mapping.source_column
            common_values.append(
                CaseDocCommonValueData(
                    key=key,
                    value=value,
                    source_table=source_table,
                    source_column=source_column,
                    source=f"{source_table}.{source_column}",
                )
            )
        return common_values

    def _resolve_device_placeholders_by_host_name(
        self,
        host_assignments: list[CaseDocHostAssignmentData],
        mappings: list[CaseDocPlaceholderMappingItemData],
    ) -> list[CaseDocResolvedPlaceholderData]:
        resolved_placeholders: list[CaseDocResolvedPlaceholderData] = []
        device_types = _unique_ordered(
            mapping.device_type or ""
            for mapping in mappings
            if mapping.enabled and mapping.scope == "device" and mapping.device_type
        )
        for device_type in device_types:
            resolved_placeholders.extend(
                _resolve_device_placeholders_from_values(
                    host_assignments,
                    self._load_device_values_by_host_name(mappings, device_type),
                    mappings,
                    device_type,
                )
            )
        return resolved_placeholders
