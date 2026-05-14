"""Generate case document workbook bytes from the case document template."""

from __future__ import annotations

from copy import copy
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from app.core.responses import CaseDocResolveContextData, ModuleRowData, SourceDocDetailData


XLSM_MEDIA_TYPE = "application/vnd.ms-excel.sheet.macroEnabled.12"
TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "templates" / "case_docs" / "case_doc_template.xlsm"
RESOLVED_VALUES_SHEET_NAME = "\u89e3\u6c7a\u5024"
SOURCE_DOC_EXPANSION_SHEET_NAME = "\u539f\u672c\u5c55\u958b"
BODY_HEADER_MARKERS = {"\u5927", "\u4e2d", "\u5c0f", "\u4f5c\u696d\u5185\u5bb9", "\u78ba\u8a8d\u4e8b\u9805 or \u9805\u76ee", "\u30b3\u30de\u30f3\u30c9"}
FOOTER_START_MARKERS = {"\u9023\u7d61\u4e8b\u9805"}
FOOTER_GAP_ROWS = 3
BODY_MAX_COLUMN = 13
BODY_HEADER_TO_FIELD = {
    "\u5927": "major_no",
    "\u4e2d": "middle_no",
    "\u5c0f": "minor_no",
    "\u6280\u8853\u8cc7\u6599\u540d": "tech_doc_text",
    "\u4f5c\u696d\u5185\u5bb9": "work_text",
    "\u78ba\u8a8d\u4e8b\u9805 or \u9805\u76ee": "expected_result",
    "\u6642\u523b": "time_text",
    "window": "window_text",
    "P": "p_text",
    "\u30b3\u30de\u30f3\u30c9": "command_text",
}
BODY_FALLBACK_COLUMNS = {
    "major_no": 1,
    "middle_no": 2,
    "minor_no": 3,
    "tech_doc_text": 4,
    "work_text": 5,
    "expected_result": 9,
    "time_text": 10,
    "window_text": 11,
    "p_text": 12,
    "command_text": 13,
}
TEMPLATE_PLACEHOLDER_ALIASES = {
    "DEVICE_NAME": "TARGET_DEVICE_HOSTNAME",
    "NW_ADDRESS": "SBC_COMMAND_FLOATING_IP",
    "USER": "LOGIN_USER",
}


def _u(value: str) -> str:
    """Decode escaped Japanese literals while keeping this source ASCII-safe."""

    try:
        return value.encode("ascii").decode("unicode_escape")
    except UnicodeEncodeError:
        return value


def _clear_sheet(sheet: Worksheet) -> None:
    """Clear cell values from a worksheet while keeping the sheet object."""

    if sheet.max_row > 1:
        sheet.delete_rows(1, sheet.max_row)
    elif sheet.max_row == 1:
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = None


def _write_row(sheet: Worksheet, row_index: int, values: list[object]) -> None:
    for column_index, value in enumerate(values, start=1):
        sheet.cell(row=row_index, column=column_index, value=value)


def _style_heading(sheet: Worksheet, row_index: int) -> None:
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in sheet[row_index]:
        cell.font = Font(bold=True)
        cell.fill = fill


def _cell_text(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _find_case_doc_body_sheet(workbook) -> Worksheet | None:
    for sheet in workbook.worksheets:
        if sheet.title in {RESOLVED_VALUES_SHEET_NAME, SOURCE_DOC_EXPANSION_SHEET_NAME}:
            continue
        if _find_body_header_row(sheet) is not None:
            return sheet
    return None


def _find_body_header_row(sheet: Worksheet) -> int | None:
    markers = {_u(marker) for marker in BODY_HEADER_MARKERS}
    for row_index in range(1, sheet.max_row + 1):
        row_values = {_cell_text(sheet.cell(row=row_index, column=column_index).value) for column_index in range(1, sheet.max_column + 1)}
        if markers.issubset(row_values):
            return row_index
    return None


def _detect_body_columns(sheet: Worksheet, header_row: int) -> dict[str, int]:
    columns = dict(BODY_FALLBACK_COLUMNS)
    header_to_field = {_u(header): field for header, field in BODY_HEADER_TO_FIELD.items()}
    for column_index in range(1, sheet.max_column + 1):
        header_text = _cell_text(sheet.cell(row=header_row, column=column_index).value)
        field_name = header_to_field.get(header_text)
        if field_name:
            columns[field_name] = column_index
    return columns


def _find_footer_start_row(sheet: Worksheet, start_row: int) -> int:
    footer_markers = {_u(marker) for marker in FOOTER_START_MARKERS}
    for row_index in range(start_row, sheet.max_row + 1):
        row_values = {_cell_text(sheet.cell(row=row_index, column=column_index).value) for column_index in range(1, sheet.max_column + 1)}
        if row_values & footer_markers:
            return row_index
    return sheet.max_row + 1


def _is_major_heading_row(module_row: ModuleRowData) -> bool:
    return bool(module_row.major_no) and str(module_row.middle_no or "") in {"", "0"} and str(module_row.minor_no or "") in {"", "0"}


def _is_footer_like_module_row(module_row: ModuleRowData) -> bool:
    footer_markers = {_u(marker) for marker in FOOTER_START_MARKERS}
    values = {
        _cell_text(module_row.tech_doc_text),
        _cell_text(module_row.work_text),
        _cell_text(module_row.expected_result),
        _cell_text(module_row.command_text),
    }
    return bool(values & footer_markers)


def _flatten_enabled_source_doc_rows(source_doc: SourceDocDetailData) -> list[ModuleRowData]:
    rows: list[ModuleRowData] = []
    for module in sorted(source_doc.items, key=lambda item: item.item_order):
        if not module.enabled:
            continue
        rows.extend(
            row
            for row in sorted(module.rows, key=lambda row: row.row_order)
            if not _is_footer_like_module_row(row)
        )
    return rows


def _snapshot_row_style(sheet: Worksheet, row_index: int) -> tuple[list[dict[str, object]], float | None]:
    styles: list[dict[str, object]] = []
    for column_index in range(1, 14):
        cell = sheet.cell(row=row_index, column=column_index)
        styles.append(
            {
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
                "protection": copy(cell.protection),
            }
        )
    return styles, sheet.row_dimensions[row_index].height


def _apply_row_style(sheet: Worksheet, row_index: int, styles: list[dict[str, object]], height: float | None) -> None:
    if height is not None:
        sheet.row_dimensions[row_index].height = height
    for column_index, style in enumerate(styles, start=1):
        cell = sheet.cell(row=row_index, column=column_index)
        cell.font = copy(style["font"])
        cell.fill = copy(style["fill"])
        cell.border = copy(style["border"])
        cell.alignment = copy(style["alignment"])
        cell.number_format = str(style["number_format"])
        cell.protection = copy(style["protection"])


def _unmerge_ranges_from_row(sheet: Worksheet, start_row: int) -> None:
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.max_row >= start_row:
            sheet.unmerge_cells(str(merged_range))


def _delete_rows_from(sheet: Worksheet, start_row: int) -> None:
    if sheet.max_row >= start_row:
        sheet.delete_rows(start_row, sheet.max_row - start_row + 1)


def _with_bottom_border(border: Border, bottom: Side) -> Border:
    return Border(
        left=copy(border.left),
        right=copy(border.right),
        top=copy(border.top),
        bottom=bottom,
        diagonal=copy(border.diagonal),
        diagonal_direction=border.diagonal_direction,
        diagonalUp=border.diagonalUp,
        diagonalDown=border.diagonalDown,
        outline=border.outline,
        vertical=copy(border.vertical),
        horizontal=copy(border.horizontal),
    )


def _apply_thick_bottom_border(sheet: Worksheet, row_index: int, max_column: int) -> None:
    thick_side = Side(style="thick", color="000000")
    for column_index in range(1, max_column + 1):
        cell = sheet.cell(row=row_index, column=column_index)
        cell.border = _with_bottom_border(cell.border, thick_side)


def _write_footer_label(sheet: Worksheet, row_index: int, styles: list[dict[str, object]], height: float | None) -> None:
    _apply_row_style(sheet, row_index, styles, height)
    for column_index in range(1, BODY_MAX_COLUMN + 1):
        sheet.cell(row=row_index, column=column_index, value=None)
    footer_cell = sheet.cell(row=row_index, column=1, value=_u("\u9023\u7d61\u4e8b\u9805"))
    footer_cell.font = copy(footer_cell.font)
    footer_cell.font = Font(
        name=footer_cell.font.name,
        sz=footer_cell.font.sz,
        b=True,
        i=footer_cell.font.i,
        color=footer_cell.font.color,
        underline=footer_cell.font.underline,
    )


def _write_source_doc_body_sheet(workbook, source_doc: SourceDocDetailData) -> None:
    body_sheet = _find_case_doc_body_sheet(workbook)
    if body_sheet is None:
        return

    header_row = _find_body_header_row(body_sheet)
    if header_row is None:
        return

    start_row = header_row + 1
    footer_start_row = _find_footer_start_row(body_sheet, start_row)
    major_style, major_height = _snapshot_row_style(body_sheet, start_row)
    normal_style, normal_height = _snapshot_row_style(body_sheet, min(start_row + 1, max(footer_start_row - 1, start_row)))
    footer_style, footer_height = _snapshot_row_style(
        body_sheet, footer_start_row if footer_start_row <= body_sheet.max_row else start_row
    )
    source_rows = _flatten_enabled_source_doc_rows(source_doc)
    body_columns = _detect_body_columns(body_sheet, header_row)

    _unmerge_ranges_from_row(body_sheet, start_row)
    _delete_rows_from(body_sheet, start_row)

    for offset, source_row in enumerate(source_rows):
        row_index = start_row + offset
        styles, height = (major_style, major_height) if _is_major_heading_row(source_row) else (normal_style, normal_height)
        _apply_row_style(body_sheet, row_index, styles, height)
        for column_index in range(1, 14):
            body_sheet.cell(row=row_index, column=column_index, value=None)
        body_sheet.cell(row=row_index, column=body_columns["major_no"], value=source_row.major_no)
        body_sheet.cell(row=row_index, column=body_columns["middle_no"], value=source_row.middle_no)
        body_sheet.cell(row=row_index, column=body_columns["minor_no"], value=source_row.minor_no)
        body_sheet.cell(row=row_index, column=body_columns["tech_doc_text"], value=source_row.tech_doc_text)
        body_sheet.cell(row=row_index, column=body_columns["work_text"], value=source_row.work_text)
        body_sheet.cell(row=row_index, column=body_columns["expected_result"], value=source_row.expected_result)
        body_sheet.cell(row=row_index, column=body_columns["time_text"], value=source_row.time_text)
        body_sheet.cell(row=row_index, column=body_columns["window_text"], value=source_row.window_text)
        body_sheet.cell(row=row_index, column=body_columns["p_text"], value=source_row.p_text)
        body_sheet.cell(row=row_index, column=body_columns["command_text"], value=source_row.command_text)

    if source_rows:
        last_source_row_index = start_row + len(source_rows) - 1
        _apply_thick_bottom_border(body_sheet, last_source_row_index, max(BODY_MAX_COLUMN, max(body_columns.values())))

    footer_row_index = start_row + len(source_rows) + FOOTER_GAP_ROWS
    _write_footer_label(body_sheet, footer_row_index, footer_style, footer_height)


def _value_for_target_host(context: CaseDocResolveContextData, placeholder: str) -> str | None:
    """Return a resolved placeholder value tied to the selected target host name."""

    target_host_name = context.target_assignment.host_name
    return next(
        (
            item.value
            for item in context.resolved_placeholders
            if item.placeholder == placeholder and item.host_name == target_host_name
        ),
        None,
    )


def _add_legacy_placeholder_aliases(values: dict[str, str]) -> None:
    for legacy_name, formal_name in TEMPLATE_PLACEHOLDER_ALIASES.items():
        if formal_name in values:
            values.setdefault(legacy_name, values[formal_name])


def _build_placeholder_values(context: CaseDocResolveContextData) -> dict[str, str]:
    """Build template values after the target host name has been resolved."""

    values = {item.key: item.value for item in context.common_values}
    values["TARGET_DEVICE_HOSTNAME"] = context.target_assignment.host_name

    target_command_floating_ip = _value_for_target_host(context, "SBC_COMMAND_FLOATING_IP")
    if target_command_floating_ip:
        values["SBC_COMMAND_FLOATING_IP"] = target_command_floating_ip

    for item in context.resolved_placeholders:
        values.setdefault(item.placeholder, item.value)

    _add_legacy_placeholder_aliases(values)
    return values


def _replace_placeholders(sheet: Worksheet, values: dict[str, str]) -> None:
    """Replace {{PLACEHOLDER}} strings in one template worksheet."""

    for row in sheet.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or "{{" not in cell.value:
                continue
            replaced = cell.value
            for key, value in values.items():
                replaced = replaced.replace(f"{{{{{key}}}}}", value)
            cell.value = replaced


def _write_resolved_values_sheet(sheet: Worksheet, context: CaseDocResolveContextData) -> None:
    """Write resolved values to an auditable worksheet in the template copy."""

    host_assignment_title = _u("\\u30db\\u30b9\\u30c8\\u5272\\u5f53")
    common_value_title = _u("\\u5171\\u901a\\u5024")
    resolved_value_title = _u("\\u89e3\\u6c7a\\u6e08\\u307f\\u5024")
    host_header = [_u("\\u30b9\\u30ed\\u30c3\\u30c8"), _u("\\u88c5\\u7f6e\\u7a2e\\u5225"), _u("\\u7cfb"), _u("\\u30db\\u30b9\\u30c8\\u540d")]
    common_header = [_u("\\u30ad\\u30fc"), _u("\\u5024"), _u("\\u51fa\\u5178")]
    resolved_header = [_u("\\u5024\\u540d"), _u("\\u5024"), _u("\\u51fa\\u5178\\u30c6\\u30fc\\u30d6\\u30eb"), _u("\\u51fa\\u5178\\u30ab\\u30e9\\u30e0"), _u("\\u30db\\u30b9\\u30c8\\u540d")]

    _clear_sheet(sheet)
    rows: list[list[object]] = [
        [_u("\\u6848\\u4ef6CS \\u751f\\u6210\\u7d50\\u679c")],
        [_u("\\u539f\\u672cID"), context.source_doc_id],
        [_u("\\u30e6\\u30cb\\u30c3\\u30c8\\u69cb\\u6210ID"), context.unit_config.unit_config_id],
        [_u("FS\\u30af\\u30e9\\u30b9\\u30bf"), context.unit_config.fs_cluster_name],
        [_u("\\u30d6\\u30ed\\u30c3\\u30af"), context.unit_config.block],
        [_u("\\u90fd\\u9053\\u5e9c\\u770c"), context.unit_config.prefecture],
        [_u("\\u30d3\\u30eb"), context.unit_config.building],
        [],
        [host_assignment_title],
        host_header,
    ]
    rows.extend(
        [assignment.slot_key, assignment.device_type, assignment.system or "-", assignment.host_name]
        for assignment in context.host_assignments
    )
    rows.extend(
        [
            [],
            [common_value_title],
            common_header,
        ]
    )
    rows.extend([item.key, item.value, item.source] for item in context.common_values)
    rows.extend(
        [
            [],
            [resolved_value_title],
            resolved_header,
        ]
    )
    rows.extend(
        [item.placeholder, item.value, item.source_table, item.source_column, item.host_name or "-"]
        for item in context.resolved_placeholders
    )

    for row_index, values in enumerate(rows, start=1):
        _write_row(sheet, row_index, values)
        if values and len(values) == 1:
            sheet.cell(row=row_index, column=1).font = Font(bold=True, size=14)
        if values in (host_header, common_header, resolved_header):
            _style_heading(sheet, row_index)

    for column_letter, width in {"A": 24, "B": 28, "C": 22, "D": 24, "E": 24}.items():
        sheet.column_dimensions[column_letter].width = width
    sheet.freeze_panes = "A10"


def _write_source_doc_expansion_sheet(sheet: Worksheet, source_doc: SourceDocDetailData) -> None:
    """Write source document modules and rows used by the generated case document."""

    source_doc_title = _u("\u539f\u672c\u5c55\u958b")
    enabled_text = _u("\u6709\u52b9")
    disabled_text = _u("\u7121\u52b9")
    row_header = [
        _u("\u30e2\u30b8\u30e5\u30fc\u30eb\u9806"),
        _u("\u6709\u52b9\u72b6\u614b"),
        _u("\u30e2\u30b8\u30e5\u30fc\u30ebID"),
        _u("\u30e2\u30b8\u30e5\u30fc\u30eb\u540d"),
        _u("\u884c\u9806"),
        _u("\u5927"),
        _u("\u4e2d"),
        _u("\u5c0f"),
        _u("\u4f5c\u696d\u5185\u5bb9"),
        _u("\u78ba\u8a8d\u7d50\u679c"),
        _u("\u30b3\u30de\u30f3\u30c9"),
    ]

    _clear_sheet(sheet)
    rows: list[list[object]] = [
        [source_doc_title],
        [_u("\u539f\u672cID"), source_doc.source_doc_id],
        [_u("\u539f\u672c\u30ad\u30fc"), source_doc.source_doc_key],
        [_u("\u539f\u672c\u540d"), source_doc.source_doc_name],
        [_u("\u7248"), source_doc.version_no],
        [_u("\u30e2\u30b8\u30e5\u30fc\u30eb\u6570"), source_doc.module_count],
        [_u("\u6709\u52b9\u30e2\u30b8\u30e5\u30fc\u30eb\u6570"), source_doc.enabled_module_count],
        [],
        [
            _u("\u203b\u6848\u4ef6CS\u751f\u6210\u3067\u306f\u3001\u539f\u672c\u306b\u7d10\u3065\u304f\u6709\u52b9\u30e2\u30b8\u30e5\u30fc\u30eb\u3092\u9806\u756a\u3069\u304a\u308a\u5c55\u958b\u3057\u307e\u3059\u3002")
        ],
        row_header,
    ]

    for module in sorted(source_doc.items, key=lambda item: item.item_order):
        module_status = enabled_text if module.enabled else disabled_text
        if not module.enabled:
            rows.append(
                [
                    module.item_order,
                    module_status,
                    module.module_key,
                    module.module_name,
                    "-",
                    "-",
                    "-",
                    "-",
                    _u("\u6848\u4ef6CS\u751f\u6210\u5bfe\u8c61\u5916"),
                    "-",
                    "-",
                ]
            )
            continue

        if not module.rows:
            rows.append(
                [
                    module.item_order,
                    module_status,
                    module.module_key,
                    module.module_name,
                    "-",
                    "-",
                    "-",
                    "-",
                    _u("\u884c\u30c7\u30fc\u30bf\u306a\u3057"),
                    "-",
                    "-",
                ]
            )
            continue

        for module_row in sorted(module.rows, key=lambda row: row.row_order):
            rows.append(
                [
                    module.item_order,
                    module_status,
                    module.module_key,
                    module.module_name,
                    module_row.row_order,
                    module_row.major_no or "",
                    module_row.middle_no or "",
                    module_row.minor_no or "",
                    module_row.work_text or "",
                    module_row.expected_result or "",
                    module_row.command_text or "",
                ]
            )

    for row_index, values in enumerate(rows, start=1):
        _write_row(sheet, row_index, values)
        if values and len(values) == 1:
            sheet.cell(row=row_index, column=1).font = Font(bold=True, size=14)
        if values == row_header:
            _style_heading(sheet, row_index)

    for column_letter, width in {
        "A": 14,
        "B": 14,
        "C": 18,
        "D": 34,
        "E": 12,
        "F": 10,
        "G": 10,
        "H": 10,
        "I": 56,
        "J": 34,
        "K": 42,
    }.items():
        sheet.column_dimensions[column_letter].width = width
    sheet.freeze_panes = "A11"


def build_case_doc_workbook_bytes(
    context: CaseDocResolveContextData,
    source_doc: SourceDocDetailData | None = None,
) -> bytes:
    """Build an xlsm workbook by copying the configured template."""

    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"case document template was not found: {TEMPLATE_PATH}")

    workbook = load_workbook(TEMPLATE_PATH, keep_vba=True)
    if source_doc is not None:
        _write_source_doc_body_sheet(workbook, source_doc)

    placeholder_values = _build_placeholder_values(context)
    for sheet in workbook.worksheets:
        if sheet.title != RESOLVED_VALUES_SHEET_NAME:
            _replace_placeholders(sheet, placeholder_values)

    if RESOLVED_VALUES_SHEET_NAME in workbook.sheetnames:
        resolved_sheet = workbook[RESOLVED_VALUES_SHEET_NAME]
    else:
        resolved_sheet = workbook.create_sheet(RESOLVED_VALUES_SHEET_NAME)
    _write_resolved_values_sheet(resolved_sheet, context)

    if source_doc is not None:
        if SOURCE_DOC_EXPANSION_SHEET_NAME in workbook.sheetnames:
            source_doc_sheet = workbook[SOURCE_DOC_EXPANSION_SHEET_NAME]
        else:
            source_doc_sheet = workbook.create_sheet(SOURCE_DOC_EXPANSION_SHEET_NAME)
        _write_source_doc_expansion_sheet(source_doc_sheet, source_doc)
        _replace_placeholders(source_doc_sheet, placeholder_values)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
